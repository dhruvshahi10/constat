"""Eval (c): nothing reaches DELIVERED without a named reviewer; the chain
survives the resumed append; exceptions can never be approved."""
from datetime import date
from pathlib import Path

import json
import pytest

from constat.models import QState
from constat.pipeline import AuditLog, run
from constat.qgen import build_questionnaire_workbook
from constat.server import config, db, review
from constat.server.demo_questions import DEMO_QUESTIONS

ROOT = Path(__file__).resolve().parents[1]
TODAY = date(2026, 8, 12)


@pytest.fixture()
def hosted_run(tmp_path, monkeypatch):
    monkeypatch.setattr(config, "DATA", tmp_path / "data")
    monkeypatch.setattr(config, "DB_PATH", tmp_path / "data" / "constat.db")
    db.init()
    out = tmp_path / "run_abc"
    qnr = build_questionnaire_workbook(DEMO_QUESTIONS, out / "run_abc.xlsx", "Acme")
    res = run(qnr, tenant="acme", evidence_root=ROOT / "data" / "evidence",
              out_dir=out, drafter_kind="mock", today=TODAY, approval_mode="human")
    from constat.report import write_report
    write_report(res, TODAY)
    return out


def test_queue_lists_cited_and_exceptions(hosted_run):
    items = review.queue_items(hosted_run)
    states = {i["question_id"]: i["state"] for i in items}
    assert "GRC_REVIEW" in states.values()
    assert "EXCEPTION" in states.values()
    cited = next(i for i in items if i["state"] == "GRC_REVIEW")
    assert cited["citations"], "cited item must show its provenance"


def test_approve_requires_name_and_note(hosted_run):
    items = review.queue_items(hosted_run)
    qid = next(i["question_id"] for i in items if i["state"] == "GRC_REVIEW")
    with pytest.raises(review.ReviewError, match="required"):
        review.act(hosted_run, qid, "approve", "", "note")
    with pytest.raises(review.ReviewError, match="required"):
        review.act(hosted_run, qid, "approve", "Priya N.", "  ")


def test_named_approval_delivers_and_chain_survives(hosted_run):
    items = review.queue_items(hosted_run)
    qid = next(i["question_id"] for i in items if i["state"] == "GRC_REVIEW")
    out = review.act(hosted_run, qid, "approve", "Priya N.", "Verified against policy.")
    assert out["state"] == "DELIVERED"

    # audit chain intact across resume, and the actor is the named human
    audit_path = hosted_run / "audit_log.jsonl"
    assert AuditLog.verify_chain(audit_path) is True
    last = json.loads(audit_path.read_text().splitlines()[-1])
    assert last["actor"] == "Priya N."
    assert last["action"] == "APPROVED"
    assert last["detail"]["mode"] == "HUMAN_REVIEW"
    assert "SIMULATED" not in json.dumps(last)

    # state.json and metrics reflect the release
    state = json.loads((hosted_run / "state.json").read_text())
    assert {q["state"] for q in state["questions"] if q["question_id"] == qid} == {"DELIVERED"}
    metrics = json.loads((hosted_run / "metrics.json").read_text())
    assert metrics["auto_approved_delivered"] == 1
    assert metrics["audit_chain_valid"] is True

    # DELIVERED.xlsx regenerated with the released answer
    from openpyxl import load_workbook
    wb = load_workbook(hosted_run / "run_abc__DELIVERED.xlsx")
    ws = wb.active
    row = next(q["row"] for q in state["questions"] if q["question_id"] == qid)
    assert ws.cell(row=row, column=4).value
    assert "[ABSTAINED]" not in str(ws.cell(row=row, column=4).value)


def test_exception_can_never_be_approved(hosted_run):
    items = review.queue_items(hosted_run)
    qid = next(i["question_id"] for i in items if i["state"] == "EXCEPTION")
    with pytest.raises(review.ReviewError, match="never be released"):
        review.act(hosted_run, qid, "approve", "Priya N.", "trying anyway")


def test_reject_returns_to_exception(hosted_run):
    items = review.queue_items(hosted_run)
    qid = next(i["question_id"] for i in items if i["state"] == "GRC_REVIEW")
    out = review.act(hosted_run, qid, "reject", "Priya N.", "Wording overclaims scope.")
    assert out["state"] == "EXCEPTION"
    state = json.loads((hosted_run / "state.json").read_text())
    assert {q["state"] for q in state["questions"] if q["question_id"] == qid} == {"EXCEPTION"}
    assert AuditLog.verify_chain(hosted_run / "audit_log.jsonl") is True
