"""Human review invariants.

Separate from `test_gates.py`, which is the constitution and is never edited.
These lock the properties that make a reviewed answer a representation rather
than a draft:

  R1  a review extends the run's own hash chain, and stays verifiable
  R2  approval cannot manufacture evidence — an unsupported answer is not approvable
  R3  a human-authored answer is labelled as such, everywhere it appears
  R4  forging who approved something breaks the chain
  R5  rejecting withdraws an answer that had already been released
  R6  decisions survive reopening the run
"""
from __future__ import annotations

import json
import shutil
import sys
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from trustops.models import AnswerStatus
from trustops.pipeline import AuditLog, run
from trustops.review import ReviewError, ReviewSession

ROOT = Path(__file__).resolve().parents[1]
EVIDENCE = ROOT / "data" / "evidence"
QNR = ROOT / "data" / "questionnaires" / "acme_security_questionnaire.xlsx"
TODAY = date(2026, 8, 8)
ACTOR = "Priya Nair <priya@acme.example>"

CITED = "IAM-01.1"          # released with a surviving citation
UNSUPPORTED = "A&A-02.1"    # refused: only a roadmap exists


@pytest.fixture()
def session(tmp_path):
    run(QNR, tenant="acme", evidence_root=EVIDENCE, out_dir=tmp_path,
        drafter_kind="mock", today=TODAY, reviewer_mode="manual")
    return ReviewSession(tmp_path)


def test_manual_mode_approves_nothing_automatically(tmp_path):
    result = run(QNR, tenant="acme", evidence_root=EVIDENCE, out_dir=tmp_path,
                 drafter_kind="mock", today=TODAY, reviewer_mode="manual")
    assert result.metrics["auto_approved_delivered"] == 0


# R1 — the review is part of the run's chain, not a second log
def test_r1_review_extends_the_same_chain(session):
    before = len((session.run_dir / "audit_log.jsonl").read_text().splitlines())
    session.decide(CITED, "approve", actor=ACTOR)
    after = (session.run_dir / "audit_log.jsonl").read_text().splitlines()
    assert len(after) == before + 1
    assert json.loads(after[-1])["actor"] == ACTOR
    assert AuditLog.verify_chain(session.run_dir / "audit_log.jsonl") is True


# R2 — approval cannot conjure evidence
def test_r2_unsupported_answer_is_not_approvable(session):
    assert session.drafts[UNSUPPORTED].status is AnswerStatus.NO_EVIDENCE
    with pytest.raises(ReviewError, match="nothing was released"):
        session.decide(UNSUPPORTED, "approve", actor=ACTOR)


def test_r2_review_requires_a_named_reviewer(session):
    with pytest.raises(ReviewError, match="named reviewer"):
        session.decide(CITED, "approve", actor="someone")


# R3 — a human-authored answer never passes as evidence-backed
def test_r3_human_authored_is_labelled_everywhere(session):
    session.decide(UNSUPPORTED, "edit", actor=ACTOR, answer="Not certified; audit planned.")
    draft = session.drafts[UNSUPPORTED]
    assert draft.status is AnswerStatus.HUMAN_AUTHORED
    assert draft.to_contract()["status"] == "human_authored"
    assert not draft.citations, "a human answer must not acquire citations it never had"

    delivered = session.export()
    ws = load_workbook(delivered).active
    row = next(q.row for q in session.questions if q.question_id == UNSUPPORTED)
    notes = str(ws.cell(row=row, column=5).value)
    assert "status=HUMAN AUTHORED" in notes
    assert "EVIDENCE BACKED" not in notes


# R4 — you cannot rewrite who signed
def test_r4_forging_the_reviewer_breaks_the_chain(session, tmp_path):
    session.decide(CITED, "approve", actor=ACTOR)
    log = session.run_dir / "audit_log.jsonl"
    assert AuditLog.verify_chain(log) is True

    forged = tmp_path / "forged.jsonl"
    shutil.copy(log, forged)
    lines = forged.read_text().splitlines()
    idx = next(i for i, ln in enumerate(lines) if json.loads(ln)["action"] == "APPROVED")
    obj = json.loads(lines[idx])
    obj["actor"] = "someone.else@acme.example"
    lines[idx] = json.dumps(obj, ensure_ascii=False)
    forged.write_text("\n".join(lines) + "\n")
    assert AuditLog.verify_chain(forged) is False


# R5 — rejection withdraws a released answer
def test_r5_reject_withdraws_the_answer(session):
    assert session.drafts[CITED].answer
    session.decide(CITED, "reject", actor=ACTOR, note="superseded policy")
    draft = session.drafts[CITED]
    assert draft.answer is None and draft.abstained
    assert draft.status is AnswerStatus.NO_EVIDENCE

    ws = load_workbook(session.export()).active
    row = next(q.row for q in session.questions if q.question_id == CITED)
    assert "[ABSTAINED]" in str(ws.cell(row=row, column=4).value)


# R6 — decisions are durable
def test_r6_decisions_survive_reopening(session):
    session.decide(CITED, "approve", actor=ACTOR, note="checked against the policy")
    reopened = ReviewSession(session.run_dir)
    decision = reopened.decisions[CITED]
    assert decision.action == "approve" and decision.actor == ACTOR
    assert reopened.drafts[CITED].requires_human is False
    assert reopened.summary()["reviewers"] == [ACTOR]
