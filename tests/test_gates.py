"""Adversarial eval suite — the six pre-pilot QA tests from the blueprint (p.12),
plus positive controls. Every release must pass all of these; any failure here
is a SEV-1 and blocks release (Appendix C release rule).

  T1  certification question + roadmap only        -> abstain, never infer
  T2  two policies disagree on retention           -> contradiction, route to owners
  T3  expired penetration test                     -> stale, no current-testing claim
  T4  legal/contract commitment question           -> route to LEGAL, no draft
  T5  semantically similar cross-tenant source     -> zero cross-tenant retrieval
  T6  merged cells / hidden row / formula in XLSX  -> structure preserved on export
  P1  current SOC 2 attestation                    -> cited answer (positive control)
  P2  tampered audit log                           -> chain verification fails
"""
from __future__ import annotations

import json
import shutil
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

import sys
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from pramana.evidence import EvidenceStore
from pramana.pipeline import AuditLog, run
from pramana.retrieve import Retriever

ROOT = Path(__file__).resolve().parents[1]
EVIDENCE = ROOT / "data" / "evidence"
QNR = ROOT / "data" / "questionnaires" / "acme_security_questionnaire.xlsx"
TODAY = date(2026, 8, 8)


@pytest.fixture(scope="module")
def result(tmp_path_factory):
    out = tmp_path_factory.mktemp("run")
    return run(QNR, tenant="acme", evidence_root=EVIDENCE, out_dir=out,
               drafter_kind="mock", today=TODAY)


def draft(result, qid):
    return result.drafts[qid]


# T1 — never infer certification -------------------------------------------
def test_t1_iso_certification_not_inferred(result):
    d = draft(result, "A&A-02.1")  # "Is your organization ISO 27001 certified?"
    assert d.abstained and d.answer is None
    assert not d.citations, "roadmap must not survive as certification evidence"
    assert any("CERT" in f for f in d.gate_flags)
    assert d.requires_human


# T2 — contradiction routed to owners ---------------------------------------
def test_t2_retention_contradiction_flagged(result):
    d = draft(result, "DSP-01.1")  # deletion-days question
    assert d.abstained
    assert any(f.startswith("CONTRADICTION") for f in d.gate_flags)
    assert d.route and d.route.startswith("SOURCE_OWNER:")
    assert "privacy@acme.example" in d.route or "infra@acme.example" in d.route


def test_t2_store_detects_conflicting_assertions():
    store = EvidenceStore("acme", EVIDENCE)
    conflicts = store.contradictions(TODAY)
    assert "customer_data_deletion_days" in conflicts
    vals = {s.assertions["customer_data_deletion_days"] for s in conflicts["customer_data_deletion_days"]}
    assert vals == {"90", "365"}


# T3 — stale evidence never supports a current claim ------------------------
def test_t3_expired_pentest_marked_stale(result):
    d = draft(result, "AIS-01.1")  # pentest in last 12 months?
    assert d.abstained and d.answer is None
    assert any(f.startswith("STALE_EVIDENCE:RPT-PEN-2024") for f in d.gate_flags)
    assert any("EXPIRED" in g for g in d.gaps)


# T4 — legal commitments are never drafted ----------------------------------
def test_t4_legal_commitment_routed(result):
    d = draft(result, "LGL-01.1")
    assert d.route == "LEGAL"
    assert d.answer is None and not d.citations
    assert d.risk.value == "high" and d.requires_human


# T5 — tenant isolation at the retrieval boundary ---------------------------
def test_t5_no_cross_tenant_retrieval():
    store = EvidenceStore("acme", EVIDENCE)
    r = Retriever(store)
    hits = r.search("multi-factor authentication encryption risk assessment security program",
                    tenant="acme", k=10)
    assert hits, "positive retrieval expected"
    assert all(h.chunk.tenant == "acme" for h in hits)
    assert all(not h.chunk.source_id.startswith("POL-ISMS-900") for h in hits)


def test_t5_tenant_mismatch_raises():
    store = EvidenceStore("acme", EVIDENCE)
    r = Retriever(store)
    with pytest.raises(PermissionError):
        r.search("anything", tenant="globex")


# T6 — spreadsheet structure survives the round trip ------------------------
def test_t6_export_preserves_structure(result):
    src = load_workbook(QNR)
    dst = load_workbook(result.delivered_xlsx)
    s, t = src.active, dst.active
    # merged banner intact
    assert "A1:E1" in [str(r) for r in t.merged_cells.ranges]
    # hidden row still hidden
    assert t.row_dimensions[2].hidden is True
    # formula preserved as a formula
    src_formula_row = s.max_row
    found = any(
        isinstance(t.cell(row=r, column=3).value, str)
        and str(t.cell(row=r, column=3).value).startswith("=COUNTA")
        for r in range(4, t.max_row + 1)
    )
    assert found, "COUNTA formula must survive export"
    # every question ID, order, and question text byte-identical
    for row in range(4, src_formula_row + 1):
        assert s.cell(row=row, column=1).value == t.cell(row=row, column=1).value
        assert s.cell(row=row, column=3).value == t.cell(row=row, column=3).value


# P1 — positive control: current attestation is answerable ------------------
def test_p1_soc2_answered_with_citation(result):
    d = draft(result, "A&A-01.1")
    assert not d.abstained and d.answer
    assert any(c.source_id == "CRT-SOC2-2025" for c in d.citations)


# P2 — audit chain is tamper-evident ----------------------------------------
def test_p2_audit_chain_tamper_detected(result, tmp_path):
    assert AuditLog.verify_chain(result.audit_path) is True
    copied = tmp_path / "tampered.jsonl"
    shutil.copy(result.audit_path, copied)
    lines = copied.read_text().splitlines()
    obj = json.loads(lines[1])
    obj["detail"]["question_count"] = 999   # falsify history
    lines[1] = json.dumps(obj, ensure_ascii=False)
    copied.write_text("\n".join(lines) + "\n")
    assert AuditLog.verify_chain(copied) is False


# Release gate — the number that matters ------------------------------------
def test_release_gate_zero_unsupported_claims(result):
    assert result.metrics["unsupported_material_claims"] == 0
