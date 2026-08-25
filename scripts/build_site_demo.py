"""Build the landing page's interactive demo data from real engine output.

Runs the deterministic mock pipeline against the synthetic acme tenant and
writes site/demo/contracts.json: one contract per showcase question, exactly
what the live engine returns, so the landing demo is truthful without an API
call or a server. Rebuild whenever gates or evidence change:

    .venv/bin/python scripts/build_site_demo.py

It also assembles site/index.html from site/index.template.html, filling:

    /*@CSS@*/       the brand stylesheet with fonts inlined
    /*@DEMO@*/      the contracts, as JSON, for the client
    <!--@CHIPS@-->  the chip buttons
    <!--@Q0@-->     the first question
    <!--@VERDICT0@-->, <!--@ANSWER0@-->, <!--@PROV0@-->, <!--@GAPS0@-->

The last five exist so the demo frame is complete on first paint. Typing the
first question a character at a time meant the hero opened on roughly two
seconds of empty box, which is the worst possible first impression for a
product whose whole claim is that it shows its work.
"""
from __future__ import annotations

import html as _html
import json
import re
import sys
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from constat.drafter import make_drafter          # noqa: E402
from constat.evidence import EvidenceStore        # noqa: E402
from constat.gates import post_gate, pre_gate     # noqa: E402
from constat.models import Draft, Question        # noqa: E402
from constat.retrieve import Retriever            # noqa: E402

# (id, label shown on the chip, question, which trap it demonstrates)
#
# The ISO question is worded to actually retrieve the certification roadmap.
# The earlier phrasing retrieved nothing at all, so the card promised "CERT
# NEVER INFERRED" while the gap text underneath said "no approved evidence
# retrieved" — a plain retrieval miss dressed up as the flagship gate. A GRC
# buyer reads gap text closely, and that contradiction is exactly the kind of
# thing they are trained to catch. This wording makes the demonstration real:
# PLN-ISO-RM is retrieved, cited, and then struck out by the cert gate for
# being a roadmap.
SHOWCASE = [
    ("mfa", "MFA enforcement", "Is multi-factor authentication enforced for all workforce access to production systems?", "cited"),
    ("ir", "Incident response", "Do you have a documented incident response plan, and within what timeframe are customers notified of a confirmed breach?", "cited"),
    ("soc2", "SOC 2 attestation", "Do you hold a current SOC 2 Type II attestation?", "cited"),
    ("iso", "ISO 27001 trap", "Is your ISMS ISO 27001 certified? Provide the certificate number, certification body and expiry date.", "cert"),
    ("del", "Deletion contradiction", "Within how many days of contract termination is customer data deleted?", "contradiction"),
    ("pen", "Stale pentest", "Has an independent penetration test been performed in the last 12 months?", "stale"),
    ("legal", "Unlimited liability", "Will Vendor contractually commit to unlimited liability for any breach?", "legal"),
]

FIXED_TODAY = date(2026, 8, 16)


def verdict_for(d) -> tuple[str, str]:
    """Label a finished draft. Order matters: most specific gate wins.

    CERT_INFERENCE_BLOCKED and a bare CERT_CLAIM are different events and used
    to share a label. The first is the flagship trap: a certification claim was
    asked, real evidence came back, and the gate struck it out for being the
    wrong evidence class. The second is a certification question that simply
    found nothing, which is an ordinary retrieval miss. Calling both "CERT
    NEVER INFERRED" put a verdict on the card that the gap text underneath
    contradicted.
    """
    if d.route == "LEGAL":
        return "ROUTED TO COUNSEL", "legal"
    if d.abstained and any(f.startswith("CONTRADICTION") for f in d.gate_flags):
        return "REFUSED, SOURCES CONFLICT", "warn"
    if d.abstained and any(f.startswith("STALE_EVIDENCE") for f in d.gate_flags):
        return "REFUSED, EVIDENCE EXPIRED", "warn"
    if d.abstained and any(f.startswith("CERT_INFERENCE_BLOCKED") for f in d.gate_flags):
        return "REFUSED, CERT NEVER INFERRED", "warn"
    if d.abstained and any(f.startswith("UNGROUNDED_ANSWER") for f in d.gate_flags):
        return "REFUSED, NOT GROUNDED IN SOURCE", "warn"
    if d.abstained and any(f.startswith("CERT_CLAIM") for f in d.gate_flags):
        return "REFUSED, NO CERTIFICATE ON FILE", "warn"
    if d.abstained:
        return "REFUSED, GAP NAMED", "warn"
    if d.gaps or d.requires_human:
        # something citable survived, but a gate raised a finding on the way:
        # honest label, not "clean"
        return "CITED, HELD FOR HUMAN REVIEW", "rev"
    return "CITED, GATE CLEAN", "ok"


def answer(question_text: str, store: EvidenceStore, retriever: Retriever) -> dict:
    q = Question(question_id="DEMO", row=0, domain="Demo", text=question_text)
    d = Draft(question_id="DEMO", answer=None)
    d = pre_gate(q, d)
    if d.route != "LEGAL":
        d = make_drafter("mock", retriever).draft(q, "acme")
        d = pre_gate(q, d)
    d = post_gate(q, d, store, FIXED_TODAY)

    verdict, tone = verdict_for(d)
    return {
        "verdict": verdict, "tone": tone, "answer": d.answer,
        "citations": [f"{c.source_id} v{c.version} {c.location}" for c in d.citations],
        "gaps": d.gaps[:2], "flags": d.gate_flags[:2],
    }


CHIP_CLASS = {"ok": "chip c-ok", "legal": "chip c-sig", "rev": "chip c-rev"}

# --- the question bank -------------------------------------------------------
# Roughly forty buyer-phrased questionnaire items, distinct from the showcase
# seven, spanning the breadth of a real vendor security questionnaire. Every
# one runs through the same answer() path as the showcase, so whatever verdict
# lands in contracts.json is what the engine actually returned. The client-side
# matcher picks a bank entry by keyword overlap, so each entry also carries a
# keyword list derived from its own wording.
BANK_QUESTIONS = [
    "Is customer data encrypted at rest, and with what algorithm and key length?",
    "Is data encrypted in transit using TLS 1.2 or higher?",
    "How are encryption keys managed, stored, and rotated?",
    "Are backups performed regularly, and are they encrypted and tested for restore?",
    "Do you maintain a documented disaster recovery plan, and how often is it tested?",
    "What are your recovery time objective (RTO) and recovery point objective (RPO)?",
    "Do you have a formal incident response policy with defined roles and severity levels?",
    "What is your breach notification window for incidents involving personal data?",
    "Describe your vulnerability management program, including scanning frequency.",
    "What are your patching SLAs for critical, high, and medium severity vulnerabilities?",
    "Who performs your penetration tests, and can you share the executive summary?",
    "Can you provide your most recent SOC 2 report under NDA?",
    "Do you hold an ISO 27001 certificate covering the services provided to us?",
    "How is access to customer data restricted and reviewed?",
    "Do you enforce least privilege and role based access control for internal systems?",
    "How quickly is access revoked when an employee leaves or changes role?",
    "Are security events centrally logged and monitored for anomalies?",
    "How long are audit logs retained, and are they protected from tampering?",
    "How do you assess and monitor the security of your vendors and subprocessors?",
    "Will you notify us before adding a new subprocessor that handles our data?",
    "How is customer data deleted at our request, and is deletion certified?",
    "What is your data retention schedule for customer content and derived data?",
    "In which regions is customer data stored, and can we choose our data residency?",
    "Does your development lifecycle include security requirements and threat modeling?",
    "Is all code peer reviewed before it is merged and deployed to production?",
    "How are secrets such as API keys and credentials stored and rotated?",
    "Do you scan third party dependencies for known vulnerabilities?",
    "Do all employees receive security awareness training, and how often?",
    "Are background checks performed on employees before they are hired?",
    "What physical security controls protect the data centers hosting our data?",
    "Are company endpoints protected with disk encryption and endpoint detection tooling?",
    "Is your production network segmented from corporate and development networks?",
    "What DDoS mitigation protections are in place for your public services?",
    "What uptime SLA do you commit to, and how are service credits handled?",
    "Will you contractually commit to a liability cap higher than twelve months of fees?",
    "Do you carry cyber liability insurance, and what is the coverage amount?",
    "How does your privacy program address GDPR and India's DPDP Act obligations?",
    "Have you appointed a data protection officer, and how can they be contacted?",
    "Do we have the right to audit your security controls or receive audit reports?",
    "Which multi-factor authentication methods are supported, and is phishing resistant MFA available?",
    "What password complexity, rotation, and storage requirements do you enforce?",
    "How are user sessions managed, including timeout and revocation on logout?",
]

# Tokens carrying no matching signal. The list from the brief plus the same
# kind of glue words that appear in the bank's own phrasing.
KEYWORD_STOPWORDS = set(
    "is are the of for your you do does a an in to and what how within with any "
    "we our us it its be can will have has had that this these those such as or "
    "on at by from under before after when which who whom whose they them their "
    "there here was were been being not new all often long quickly higher please "
    "describe provide including involving performed".split())

# Where a question's main term has an obvious synonym a visitor might type,
# carry it so the client-side matcher hits. Applied per token, values appended.
KEYWORD_SYNONYMS = {
    "multi-factor": ["mfa", "2fa", "multifactor"],
    "mfa": ["multi-factor", "2fa"],
    "penetration": ["pentest"],
    "pentest": ["penetration"],
    "backups": ["backup"],
    "backup": ["backups"],
    "encrypted": ["encryption"],
    "encryption": ["encrypted"],
    "subprocessor": ["subprocessors", "vendor"],
    "subprocessors": ["subprocessor", "vendors"],
    "disaster": ["dr"],
    "logged": ["logging", "logs"],
    "logs": ["logging"],
    "slas": ["sla"],
    "sla": ["slas"],
    "ddos": ["denial-of-service"],
    "iso": ["27001"],
    "27001": ["iso"],
    "soc": ["soc2"],
    "gdpr": ["privacy"],
    "dpdp": ["privacy"],
    "uptime": ["availability"],
    "deleted": ["deletion"],
    "deletion": ["deleted"],
    "secrets": ["credentials"],
    "credentials": ["secrets"],
    "endpoints": ["endpoint", "laptops"],
    "officer": ["dpo"],
    "dpo": ["officer"],
    "vulnerabilities": ["vulnerability"],
    "vulnerability": ["vulnerabilities"],
    "patching": ["patch", "patches"],
    "training": ["awareness"],
    "retention": ["retained"],
    "retained": ["retention"],
    "rto": ["recovery"],
    "rpo": ["recovery"],
    "tested": ["testing"],
    "tests": ["testing", "test"],
    "insurance": ["insured"],
    "segmented": ["segmentation"],
    "liability": ["liable"],
}

_TOKEN_RE = re.compile(r"[a-z0-9]+(?:[-/.][a-z0-9]+)*")


def keywords_for(text: str) -> list[str]:
    """Deduplicated lowercased content tokens plus synonyms, for the matcher."""
    out, seen = [], set()
    for tok in _TOKEN_RE.findall(text.lower()):
        if tok in KEYWORD_STOPWORDS:
            continue
        for k in (tok, *KEYWORD_SYNONYMS.get(tok, ())):
            if k not in seen:
                seen.add(k)
                out.append(k)
    return out


def build_bank(store: EvidenceStore, retriever: Retriever) -> list[dict]:
    """Run every bank question through the real engine and keep its verdict."""
    bank, dist = [], {}
    for text in BANK_QUESTIONS:
        a = answer(text, store, retriever)
        dist[a["verdict"]] = dist.get(a["verdict"], 0) + 1
        bank.append({
            "question": text, "verdict": a["verdict"], "tone": a["tone"],
            "answer": a["answer"], "citations": a["citations"],
            "gaps": a["gaps"], "keywords": keywords_for(text),
        })
    print(f"bank: {len(bank)} questions, verdict distribution:")
    for verdict, n in sorted(dist.items(), key=lambda kv: -kv[1]):
        print(f"  {n:3}  {verdict}")
    return bank


# --- the audit chain sample --------------------------------------------------
# The live tamper widget lets a visitor edit an event field in the browser and
# recompute the chain, so it needs genuine events and the exact hashing recipe
# AuditLog uses: sha256 over json.dumps(event, sort_keys=True) with Python's
# default separators, where event = {ts, actor, action, subject, detail,
# prev_hash}. The events come from a real pipeline run, not from hand-written
# fixtures, and every hash is recomputed here before it is shipped, so a drift
# in the recipe fails the build instead of shipping a widget that lies.

CHAIN_RECIPE = ("sha256 over json.dumps(event, sort_keys=True) with ', '/': ' "
                "separators; event={ts,actor,action,subject,detail,prev_hash}")
_CHAIN_FIELDS = ("ts", "actor", "action", "subject", "detail", "prev_hash")
CHAIN_EVENTS = 6


def _chain_hash(event: dict) -> str:
    import hashlib
    body = {k: event[k] for k in _CHAIN_FIELDS}
    payload = json.dumps(body, sort_keys=True, ensure_ascii=False)
    return hashlib.sha256(payload.encode()).hexdigest()


def build_chain(root: Path) -> dict:
    import tempfile

    from constat.pipeline import GENESIS, run as pipeline_run

    with tempfile.TemporaryDirectory(prefix="constat-sitechain-") as td:
        result = pipeline_run(
            root / "data" / "questionnaires" / "acme_security_questionnaire.xlsx",
            tenant="acme", evidence_root=root / "data" / "evidence",
            out_dir=Path(td), drafter_kind="mock", today=FIXED_TODAY)
        events = [json.loads(ln) for ln
                  in result.audit_path.read_text(encoding="utf-8").splitlines()
                  if ln.strip()]

    # The browser canonicalizer assumes JSON.stringify-compatible escaping, so
    # every string in the shipped window must be pure ASCII. Prefer the start
    # of the log (prev of the first event is then GENESIS); slide forward only
    # if a window contains non-ASCII text.
    def _is_ascii(ev: dict) -> bool:
        return json.dumps(ev, ensure_ascii=False).isascii()

    window = None
    for i in range(len(events) - CHAIN_EVENTS + 1):
        if all(_is_ascii(events[i + j]) for j in range(CHAIN_EVENTS)):
            window = events[i:i + CHAIN_EVENTS]
            break
    if window is None:
        raise SystemExit("chain: no contiguous ASCII-clean window of "
                         f"{CHAIN_EVENTS} events in {len(events)} audit events")

    # Permanent recipe assertion: if AuditLog's hashing ever drifts from
    # CHAIN_RECIPE, this build must fail rather than ship a broken widget.
    for j, ev in enumerate(window):
        recomputed = _chain_hash(ev)
        if recomputed != ev["hash"]:
            raise SystemExit(
                f"chain: event {j} hash mismatch, recipe has drifted from "
                f"AuditLog (stored {ev['hash'][:12]}, recomputed {recomputed[:12]})")
        if j and ev["prev_hash"] != window[j - 1]["hash"]:
            raise SystemExit(f"chain: event {j} does not link to event {j - 1}")

    return {
        "recipe": CHAIN_RECIPE,
        "genesis": GENESIS,
        "events": [{k: ev[k] for k in (*_CHAIN_FIELDS, "hash")} for ev in window],
    }

# --- legal pages -------------------------------------------------------------
# site/legal/*.html are built here rather than hand-maintained, for the same
# reason the landing page is: they use brand.stylesheet(), so they cannot drift
# from the product's palette, type or focus rules. Jurisdiction and contact are
# module constants below rather than placeholders; both were resolved from the
# founder's own portfolio rather than asked of him again.

# The operator's real jurisdiction. It was a placeholder for one round because I
# asked the founder a legal-sounding question he had no way to answer; it is in
# fact a fact, stated in his own portfolio (Bengaluru and Delhi NCR).
JURISDICTION = "India"
LINKEDIN = "https://www.linkedin.com/in/dhruvshahi-/"
# A data subject exercising a DPDP or GDPR right cannot be told their only
# channel is a social DM, so the rights address stays an email.
RIGHTS_EMAIL = "dhruv.shahi07@gmail.com"

LEGAL_CSS = """
body{padding-bottom:80px}
.legalwrap{max-width:760px}
.draftnote{background:var(--status-warn-wash);border:1px solid var(--line-1);
border-left:2px solid var(--status-warn);border-radius:var(--radius-2);padding:16px 18px;
margin:26px 0;font-size:13.5px;line-height:1.65;color:var(--text-primary)}
.draftnote b{color:var(--status-warn);font-weight:400;font-family:var(--font-mono);
font-size:10.5px;letter-spacing:0.13em;text-transform:uppercase;display:block;margin-bottom:7px}
.ph{font-family:var(--font-mono);font-size:0.92em;color:var(--status-warn);
background:var(--status-warn-wash);padding:1px 5px;border-radius:var(--radius-1)}
/* scoped: bare h2/h3 rules here leaked 44px and 26px of top margin onto
   every card heading of the static landing page when this block is inlined */
.legalwrap h2{margin-top:44px}
.legalwrap h3{font-size:18px;margin-top:26px;margin-bottom:6px}
.legalwrap p{font-size:14.5px;line-height:1.72;color:var(--text-secondary);margin-top:10px;
max-width:74ch}
.legalwrap li{font-size:14.5px;line-height:1.72;color:var(--text-secondary);margin-top:8px}
.legalwrap ul{margin:8px 0 0 20px}
.legalwrap strong{color:var(--text-primary);font-weight:500}
.tbl{margin-top:16px;overflow-x:auto}
.tbl table{min-width:520px}
.updated{font-family:var(--font-mono);font-size:10.5px;letter-spacing:0.13em;
text-transform:uppercase;color:var(--text-tertiary);margin-top:14px}
.navback{margin-top:26px;font-family:var(--font-mono);font-size:11px}
"""

DRAFT_NOTE = """<div class="draftnote"><b>Founder drafted, pending legal review</b>
This document was written by the founder, not by a lawyer, and has not been through legal
review. It is published in this state deliberately: a pre-launch demo that collects documents
should say what it does with them from the first day rather than from the day it can afford
counsel. Where it is wrong it will be corrected, and material changes will be dated here.
If anything below matters to your decision, ask and you will get a straight answer in
writing.</div>"""


def _legal_page(title: str, heading: str, eyebrow: str, body: str) -> str:
    from constat import brand
    return (
        '<!doctype html><html lang="en"><head><meta charset="utf-8">'
        '<meta name="viewport" content="width=device-width,initial-scale=1">'
        f"<title>{title}</title>"
        f"<style>{brand.stylesheet(LEGAL_CSS)}</style></head>"
        '<body><div class="wrap legalwrap">'
        f'<header><div class="eyebrow"><b>Constat</b> / {eyebrow}</div>'
        f"<h1>{heading}</h1>"
        '<div class="updated">Version 0.1 / 16 August 2026 / operator: Constat, '
        "operated by Dhruv Shahi</div></header>"
        f"{DRAFT_NOTE}{body}"
        '<p class="navback"><a href="/">Back to Constat</a> / '
        '<a href="/site/legal/terms.html">Terms</a> / '
        '<a href="/site/legal/privacy.html">Privacy and data handling</a></p>'
        "<footer>Constat, operated by Dhruv Shahi / "
        f"governing law and venue: {JURISDICTION} / "
        f'<a href="{LINKEDIN}" target="_blank" rel="noopener">message on LinkedIn</a>'
        f' / <a href="mailto:{RIGHTS_EMAIL}">{RIGHTS_EMAIL}</a></footer>'
        "</div></body></html>")


TERMS_BODY = """
<h2>1. Who you are contracting with</h2>
<p>The service at this address is <strong>Constat, operated by Dhruv Shahi</strong>, an
individual operator based in India. There is no company entity behind it yet. Governing law and
the venue for any dispute are the laws of <strong>India</strong>. This document was drafted by the
operator rather than by counsel, so if you need contractual certainty, ask before you rely on
this service, and expect a reviewed agreement before any paid tier is sold to you.</p>

<h2>2. This is a demo, and it is sold as one</h2>
<p>Constat is pre-launch. It is provided <strong>as is and as available</strong>, with
<strong>no warranty</strong> of any kind, express or implied, including any warranty of
merchantability, fitness for a particular purpose, accuracy or non infringement. There is
<strong>no service level agreement</strong>, no uptime commitment, no support commitment and no
guarantee of continuity. The service may change, break or be withdrawn without notice, and a
workspace may be deleted early if the demo is being abused or if it costs more to run than the
operator can carry.</p>
<p>Nothing this service outputs is legal, audit or compliance advice. Every answer it drafts is
a draft. The whole design of the product is that a named human reviews before anything is
released, and that design assumes you actually do the reviewing. If you send a Constat answer
to a customer without reading it, that is your representation to your customer, not ours.</p>

<h2>3. What you may upload</h2>
<p>On the free demo tier, drafting runs on <strong>Google Gemini's free tier</strong>, whose
terms permit Google to use submitted content to improve their services. So the rule for this
tier is simple and you agree to it by uploading:</p>
<ul>
<li><strong>Upload public evidence only.</strong> A published SOC 2 report, an ISO certificate,
trust-center policies, anything you would be comfortable seeing outside your company.</li>
<li><strong>Do not upload confidential, personal or regulated data.</strong> No customer data,
no personal data of third parties, no secrets, no material non-public information.</li>
<li>Do not upload anything you do not have the right to upload.</li>
<li>Do not use the service to break the law, to attack it or anyone else, or to generate
misleading claims about your security posture.</li>
</ul>
<p>If you need to run Constat over confidential evidence, that is the BYOK tier, where you
supply your own provider key and your text travels under your contract with that provider. Ask
for it. It is not built yet, and saying so is more useful to you than a checkbox that pretends
otherwise.</p>

<h2>4. Your material stays yours</h2>
<p>You keep all right, title and interest in the documents you upload and in the answers
produced from them. <strong>We claim no ownership of your intellectual property and no licence
to it beyond what is technically required to run the service you asked for</strong>: storing the
document in your workspace, indexing it so it can be retrieved, sending a question and the
retrieved excerpts to the drafting model, and generating your report and workbook. We do not
sell, licence or share your documents. We do not use them to train anything of ours; note that
this is a statement about us and not about Google, whose free-tier terms are described above and
on the <a href="/site/legal/privacy.html">privacy page</a>.</p>

<h2>5. Limits on the free tier</h2>
<ul>
<li>Three runs per workspace.</li>
<li>Twenty documents per workspace, five megabytes per file.</li>
<li>Three signups per network per day.</li>
<li>The workspace and everything in it is hard deleted fourteen days after signup.</li>
</ul>
<p>Your workspace link is the only credential. There is no password and no reset. If you lose
the link you lose the workspace, and we cannot recover it for you, because the link is the
capability and we store only a hash of it.</p>

<h2>6. Liability</h2>
<p>To the maximum extent the law allows, the operator is not liable for indirect, incidental,
special, consequential or punitive damages, or for lost profits, lost revenue, lost data or
business interruption, arising from your use of this demo. Total aggregate liability for any
claim relating to the free tier is limited to the amount you paid for it, which is nothing.
Some jurisdictions do not allow these exclusions, in which case they apply only as far as that
jurisdiction permits.</p>

<h2>7. Ending it</h2>
<p>You can stop using the service at any time, and you can ask for your workspace to be deleted
immediately rather than waiting for the fourteen day sweep. The operator may suspend or delete a
workspace that breaks section 3, that threatens the service, or that makes the demo
unaffordable to run.</p>

<h2>8. Changes and contact</h2>
<p>These terms will change, because the product is early. Material changes will be dated at the
top of this page. The fastest way to reach the operator is a
<a href="https://www.linkedin.com/in/dhruvshahi-/" target="_blank" rel="noopener">LinkedIn
message</a>. Formal requests, including corrections and deletions, can also go to
<a href="mailto:dhruv.shahi07@gmail.com">dhruv.shahi07@gmail.com</a>, and are answered by a
person.</p>
"""

PRIVACY_BODY = """
<p>Short version, because the long version below only exists so this one is checkable: we
collect your organization name, an email address, your IP address and whatever text is in the
documents you upload. We keep uploads for fourteen days and then hard delete them. Two
sub-processors, Google and Render, and no others. One cookie, which logs you in and does
nothing else. No analytics, no trackers, no advertising pixels, on any page of this site.</p>

<h2 id="data-handling">Data handling</h2>

<h3>What we collect, and why</h3>
<div class="tbl"><table>
<tr><th>What</th><th>Why</th><th>How long</th></tr>
<tr><td><strong>Organization name</strong></td><td>Names your workspace and titles the
generated questionnaire and report.</td><td>Kept after workspace deletion, in the signup
record.</td></tr>
<tr><td><strong>Email address</strong></td><td>Identifies the signup, lets us contact you about
the demo, and is offered as the default document owner on the upload form. We do not send the
workspace link by email and we do not run a mailing list.</td><td>Kept after workspace deletion,
in the signup record.</td></tr>
<tr><td><strong>IP address</strong></td><td>Rate limiting only. Three signups per network per
day is the only thing standing between this demo and an unpaid API bill.</td><td>Deleted after
seven days.</td></tr>
<tr><td><strong>Uploaded documents and their extracted text</strong></td><td>Retrieval and
drafting. This is the product.</td><td><strong>Hard deleted fourteen days after
signup</strong>, with the whole workspace directory.</td></tr>
<tr><td><strong>Run outputs</strong> (report, workbook, audit log)</td><td>The deliverables you
came for.</td><td>Deleted with the workspace at fourteen days; older run directories are pruned
sooner, keeping the newest five.</td></tr>
<tr><td><strong>Reviewer name and note</strong></td><td>Named review is the point of the audit
chain. You type the name yourself; in the demo it is self attested and unverified.</td>
<td>Stored in the approvals record, which is not swept with the workspace.</td></tr>
<tr><td><strong>Run metadata</strong> (identifiers, timings, counts, error text)</td>
<td>Operating the queue and debugging failures.</td><td>Retained.</td></tr>
<tr><td><strong>Server logs</strong></td><td>Written to the host's log stream for operations.
They contain request lines and client addresses.</td><td>Held by the host under its own
retention.</td></tr>
</table></div>
<p>We do not ask for and do not want special category personal data, payment details or
credentials. If you upload them by mistake, email us and we will delete the workspace
immediately rather than waiting for the sweep.</p>

<h3>The free tier reality, in one sentence</h3>
<p><strong>Drafting on this demo runs on Google Gemini's free tier, whose terms allow Google to
use submitted content to improve their services, which is why this tier asks you to upload
public evidence only and why we make no no-training claim on your behalf.</strong></p>
<p>We used to write "never used for training" on the landing page. It was true of us and
misleading about the stack, so it is gone. What is accurate: <em>we</em> do not train on your
documents, and on this tier we cannot make that promise for the model provider. The BYOK tier
exists so that promise comes from the provider you already have a contract with, and the Managed
tier exists so it comes from a paid tier that carries it in writing.</p>

<h3>Sub-processors</h3>
<div class="tbl"><table>
<tr><th>Who</th><th>What they do</th><th>What they receive</th></tr>
<tr><td><strong>Google</strong> (Gemini API)</td><td>Drafts each answer.</td><td>One question
and the excerpts retrieved for it, per answer. Never your full documents, never your document
list, never another tenant's material.</td></tr>
<tr><td><strong>Render</strong></td><td>Hosts the server and the disk your workspace sits
on.</td><td>Everything stored, as any host does, plus request logs.</td></tr>
</table></div>
<p><strong>There are no other sub-processors.</strong> No analytics vendor, no error tracking
vendor, no CDN, no email service provider, no third party retrieval or indexing service. If that
list ever grows, it grows on this page first.</p>

<h3>Where processing happens</h3>
<p>Retrieval, indexing, the gates, the audit chain and every generated file are produced on our
own server. The only outbound call in a run is the drafting call to Google. Tenants are
structurally isolated: each workspace has its own directory and its own index, and the retrieval
layer is rooted inside it rather than filtered after the fact.</p>

<h3>Cookies</h3>
<p>One cookie, <span class="ph">tt_&lt;workspace&gt;</span>, set the first time you open your
workspace link. It is strictly necessary: it is what keeps you signed in to that workspace. It
is HttpOnly, SameSite=Lax, Secure, scoped to your workspace path, and expires after fourteen
days. <strong>There are no analytics cookies, no advertising cookies and no third party scripts
of any kind on this site</strong>, which is why you were never shown a cookie banner. That is a
deliberate product decision and a difference worth noticing when you compare this to the other
tools in this category.</p>

<h3>Your rights</h3>
<p>Whatever framework applies to you, the practical answer is the same and does not require a
form:</p>
<ul>
<li><strong>Access</strong>: ask and we will tell you exactly what is stored against your
workspace.</li>
<li><strong>Deletion</strong>: ask and the workspace and all its files are deleted immediately.
Otherwise it happens automatically at fourteen days.</li>
<li><strong>Correction</strong>: ask, or just re-upload.</li>
<li><strong>Objection and portability</strong>: your documents are your documents; everything
the run produced is downloadable from your workspace while it exists.</li>
<li><strong>Complaint</strong>: the operator is in India, so the Digital Personal Data
Protection Act 2023 governs and the Data Protection Board of India is the body to complain to.
If you are in the EU or UK, the GDPR rights above apply to you as well and you may complain to
your own national supervisory authority instead.</li>
</ul>
<p>Requests go to <a href="mailto:dhruv.shahi07@gmail.com">dhruv.shahi07@gmail.com</a>, or by
<a href="https://www.linkedin.com/in/dhruvshahi-/" target="_blank" rel="noopener">LinkedIn message</a> if that is easier. They
are read by one person, which in practice means a same week answer, not a ticket number.</p>

<h3>Children</h3>
<p>This is a business tool and is not directed at anyone under 16.</p>

<h3>Security, honestly</h3>
<p>Your workspace link is a bearer capability: anyone holding it holds the workspace. We store
only a hash of it. There is no password, no multi-factor authentication and no session
management beyond that cookie, because this is a demo. Treat the link like a password, and do
not put evidence in this tier that would hurt you if the link leaked.</p>
"""


def build_legal(root: Path) -> list[Path]:
    out_dir = root / "site" / "legal"
    out_dir.mkdir(parents=True, exist_ok=True)
    written = []
    for name, title, heading, eyebrow, body in (
        ("terms.html", "Constat terms of use", "Terms of use", "legal / terms", TERMS_BODY),
        ("privacy.html", "Constat privacy and data handling",
         "Privacy and data handling", "legal / privacy", PRIVACY_BODY),
    ):
        p = out_dir / name
        p.write_text(_legal_page(title, heading, eyebrow, body), encoding="utf-8")
        written.append(p)
    return written


def prerender(first: dict) -> dict[str, str]:
    """Server-side render of the first contract, so the frame has content
    before a single line of demo JS runs."""
    e = _html.escape
    prov = (f'<div class="d-prov">{"<br>".join(e(c) for c in first["citations"])}</div>'
            if first["citations"] else '<div class="d-prov warn">no citation released</div>')
    gaps = "".join(f'<div class="d-gap">&#9656; {e(g)}</div>' for g in first["gaps"])
    answer_html = (f'{e(first["answer"])}' if first["answer"]
                   else "<em>No answer released.</em>")
    chip = CHIP_CLASS.get(first["tone"], "chip c-warn")
    return {
        "Q0": e(first["question"]),
        "VERDICT0": f'<span class="{chip}" id="verdict">{e(first["verdict"])}</span>',
        "ANSWER0": answer_html,
        "PROV0": prov,
        "GAPS0": gaps,
    }


def chips_html(out: dict) -> str:
    e = _html.escape
    return "".join(
        f'<button type="button" data-k="{e(k)}" aria-pressed="false">{e(v["label"])}</button>'
        for k, v in out.items())


def ticker_html(out: dict) -> str:
    """The verdict ledger, built at build time so it survives with JS off.

    The track is the same row twice because the marquee translates by -50%;
    one copy would snap back visibly at the wrap.
    """
    e = _html.escape
    items = "".join(
        f'<div class="litem" data-t="{e(v["tone"])}">'
        f'<span class="lv">{e(v["verdict"])}</span>'
        f'<span class="lq">{e(v["label"])}</span></div>'
        for v in out.values())
    row = f'<div class="lrow">{items}</div>'
    return row + row


# Product screenshots, taken from a real hosted run and inlined as data URIs so
# the page stays one self-contained file with zero external requests. Missing
# files are not fatal: the build says so and the frames render empty rather than
# breaking, because a stale screenshot is worse than an absent one.
SHOTS = {"SHOT_REPORT": "report", "SHOT_REVIEW": "review", "SHOT_WORKSPACE": "workspace"}


def media_uris(*, inline: bool) -> dict[str, str]:
    """The screen capture and its poster, inlined like the screenshots.

    The video is real footage of a complete run (scripts/capture_demo_video.py
    regenerates it). Missing files degrade to an empty src, which renders a
    blank frame rather than breaking the page; the build says so out loud.
    """
    import base64 as _b64
    # Preference order matters. demo_edit.* is the produced cut (title cards,
    # punch-ins, captions) from scripts/produce_demo_video.py; demo.webm is the
    # raw capture it was cut from. H.264 first, because iOS Safari only plays
    # VP8 WebM from 17.4 and the founder reviews on an iPhone.
    # The WebM twin is gone. Every shipping browser decodes the H.264 first
    # source, so it was 1.83MB, 30 percent of the page, for a fallback nobody
    # reached, and it cost 296ms of domInteractive on a throttled CPU.
    candidates = {
        "VIDEO_MP4": [("site/img/demo_edit.mp4", "video/mp4")],
        "POSTER_URI": [("site/img/demo_edit_poster.jpg", "image/jpeg"),
                       ("site/img/demo_poster.jpg", "image/jpeg")],
    }
    out = {}
    for token, opts in candidates.items():
        for rel, mime in opts:
            path = ROOT / rel
            if path.is_file():
                kb = path.stat().st_size // 1024
                # The film is externalised on the hosted build and inlined only
                # where the page must be one file. Inlining it is what made the
                # document megabytes long, which is what pushed the script to
                # 99 percent of the stream, and it is also the sole reason the
                # CSP needed a media-src data: exception.
                if inline or not rel.endswith(".mp4"):
                    b64 = _b64.b64encode(path.read_bytes()).decode("ascii")
                    out[token] = f"data:{mime};base64,{b64}"
                    print(f"  {token}: {rel} ({kb}KB, inlined)")
                else:
                    out[token] = "/" + rel
                    print(f"  {token}: {rel} ({kb}KB, linked)")
                break
        else:
            print(f"  WARNING: no file for {token}, the film frame will be blank")
            out[token] = ""
    return out


def chain_html(chain: dict) -> str:
    """Prerendered tamper-widget blocks, so the chain is visible with JS off.

    Inputs render readonly; the widget's JS removes the attribute when it can
    actually recompute hashes. The hash shown is the stored one, truncated to
    the same 18 characters the live recompute displays.
    """
    e = _html.escape
    # Display labels only: the event data (and therefore every hash) keeps the
    # engine's own action names verbatim; the raw name rides in a title
    # attribute. CLASSIFIED as visible text trips the brand's dossier-language
    # rule, and the brand rule wins on a marketing page.
    labels = {"CLASSIFIED": "ROUTED", "EVIDENCE_INTEGRITY_SCAN": "EVIDENCE SCAN"}
    # The hash is rendered ON the connector rather than inside the block,
    # because that is literally what it is: the link from one event to the
    # next. Six boxes each holding a hex string read as six unrelated cards;
    # a chain whose links carry the hashes reads as a chain.
    out = []
    for i, ev in enumerate(chain["events"]):
        label = labels.get(ev["action"], ev["action"].replace("_", " "))
        if i:
            out.append(f'<span class="tlink" data-i="{i}" aria-hidden="true">'
                       f'<i></i><em>{e(chain["events"][i - 1]["hash"][:6])}</em></span>')
        out.append(
            f'<div class="tb" data-i="{i}">'
            f'<span class="th">{i + 1:02d} <s>/</s> {e(label)}</span>'
            f'<input data-f="actor" value="{e(ev["actor"])}" '
            f'aria-label="actor of event {i + 1}, edit to tamper with the chain">'
            f'<span class="hash">{e(ev["hash"][:14])}</span></div>')
    return "".join(out)


# --- the workbook proof ------------------------------------------------------
# The page's job is to be understood in five seconds by someone who has never
# heard the word "gate". The fastest way to do that is to show the artefact the
# reader already has a mental model of: the questionnaire spreadsheet, with the
# last column filled in. These rows are read out of the real DELIVERED workbook
# at build time, so the marketing page cannot drift from the engine's output.
#
# Five rows chosen to show every outcome the engine can produce: two cited, one
# refused for expired evidence, one refused for contradicting sources, and one
# routed to counsel without ever reaching the model.
WORKBOOK_FILE = ROOT / "examples" / "acme_security_questionnaire__DELIVERED.xlsx"
WORKBOOK_ROWS = ["GRC-01.1", "A&A-01.1", "AIS-01.1", "DSP-01.1", "LGL-01.1"]


def _dedash(s: str) -> str:
    """Engine text is verbatim except for dashes.

    The drafter writes em dashes ("Insufficient approved evidence - see
    exception notes"), the brand copy gate bans them in rendered copy, and the
    engine's own wording is not going to be changed to suit a web page. So the
    dash normalizes to a comma here, at the rendering boundary only.
    """
    return s.replace("—", ",").replace("–", ",").replace(" ,", ",")


def _clip(s: str, n: int) -> str:
    """Clip to n characters, preferring a natural end over a hard cut.

    Questionnaire items are often two sentences ("Do you hold a current SOC 2
    Type II attestation? If yes, state the criteria covered"). Cutting at the
    first question mark reads as the whole question; cutting at a word boundary
    mid-clause reads as broken text.
    """
    s = " ".join(s.split())
    if len(s) <= n:
        return s
    # a complete question reads better than a truncation, so allow a little
    # overshoot to reach the question mark
    q = s.find("?")
    if 0 < q <= int(n * 1.4):
        return s[:q + 1]
    return s[:n].rsplit(" ", 1)[0].rstrip(" ,;:.") + "..."


def workbook_rowcount() -> int:
    """Data rows in the real questionnaire, for the sheet header.

    Hardcoding a number here would be the one dishonest pixel on a page whose
    entire argument is that it does not guess.
    """
    try:
        import openpyxl
    except ImportError:
        return 0
    if not WORKBOOK_FILE.is_file():
        return 0
    ws = openpyxl.load_workbook(WORKBOOK_FILE).active
    # a real question id, not the hidden metadata row whose "internal-use"
    # also contains a dash
    qid = re.compile(r"^[A-Z&]{2,5}-\d+\.\d+$")
    return sum(1 for r in range(1, ws.max_row + 1)
               if isinstance(ws.cell(r, 1).value, str)
               and qid.match(ws.cell(r, 1).value.strip())
               and ws.cell(r, 3).value)


def workbook_html() -> str:
    """Render the real delivered rows as a sheet.

    Two rules learned from an audit. First, nothing is truncated with a literal
    ellipsis here: the full engine text goes into the markup and CSS clamps it,
    so a questionnaire question never renders as placeholder-looking copy.
    Second, a refused row shows its GAP rather than the generic abstain
    sentence. The engine returns byte-identical response text for every
    abstention ("Insufficient approved evidence, see exception notes"), so
    printing that put the same sentence on two consecutive rows and hid the
    only thing that differs: one document expired, the other two conflict.
    """
    try:
        import openpyxl
    except ImportError:
        print("  WARNING: openpyxl missing, the workbook proof will be empty")
        return ""
    if not WORKBOOK_FILE.is_file():
        print(f"  WARNING: {WORKBOOK_FILE.name} missing, the workbook proof will be empty")
        return ""

    ws = openpyxl.load_workbook(WORKBOOK_FILE).active
    by_id = {}
    for r in range(1, ws.max_row + 1):
        qid = ws.cell(r, 1).value
        if qid in WORKBOOK_ROWS:
            by_id[qid] = (r, ws.cell(r, 3).value or "", ws.cell(r, 4).value or "",
                          ws.cell(r, 5).value or "")

    e = _html.escape
    out = []
    for qid in WORKBOOK_ROWS:
        if qid not in by_id:
            print(f"  WARNING: row {qid} not in the workbook, skipped")
            continue
        rownum, question, response, notes = by_id[qid]
        question, response, notes = (_dedash(str(v)) for v in (question, response, notes))

        gaps = ""
        for raw in notes.splitlines():
            if raw.startswith("Gaps:"):
                gaps = raw[5:].strip()
                break
        evidence = ""
        for raw in notes.splitlines():
            if raw.startswith("Evidence:"):
                evidence = raw.replace("Evidence:", "").strip()
                break

        if response.startswith("[ROUTED TO LEGAL]"):
            tone, verdict = "legal", "Routed to counsel"
            body, tail = _split_gap(gaps)
            tail = tail or "Routed to legal"
        elif response.startswith("[ABSTAINED]"):
            tone, verdict = "warn", "Refused"
            body, tail = _split_gap(gaps)
        else:
            tone, verdict = "ok", "Cited"
            body, tail = response, evidence

        out.append(
            f'<div class="wbrow" data-tone="{tone}">'
            f'<span class="wbn" aria-hidden="true">{rownum}</span>'
            f'<span class="wbid">{e(qid)}</span>'
            f'<p class="wbq">{e(question)}</p>'
            f'<div class="wbr"><span class="wbv">{e(verdict)}</span>'
            f'<p>{e(body)}</p>'
            + (f'<span class="wbe">{e(tail)}</span>' if tail else "")
            + "</div></div>")
    return "".join(out)


def _split_gap(gaps: str) -> tuple[str, str]:
    """Split a gap into the finding and the routing instruction.

    "RPT-PEN-2024 v1.0: EXPIRED 2025-06-10, cannot support a current-state
    claim; route to security@acme.example" becomes the finding and the route,
    so the cell leads with what is wrong and closes with who owns it.
    """
    first = gaps.split("|")[0].strip()
    if ";" in first:
        head, _, tail = first.partition(";")
        return head.strip().rstrip(","), tail.strip()
    if ", route to " in first:
        head, _, tail = first.partition(", route to ")
        return head.strip(), "route to " + tail.strip()
    return first, ""


def shot_uris() -> dict[str, str]:
    path = ROOT / "site" / "img" / "shots.json"
    if not path.is_file():
        print("  WARNING: site/img/shots.json missing, product frames will be blank")
        return {k: "" for k in SHOTS}
    blobs = json.loads(path.read_text(encoding="utf-8"))
    out = {}
    for token, key in SHOTS.items():
        b64 = blobs.get(key)
        if not b64:
            print(f"  WARNING: no screenshot for {key!r}")
            out[token] = ""
        else:
            out[token] = f"data:image/jpeg;base64,{b64}"
    return out



# --- static, single-file variant ---------------------------------------------
# The hosted page depends on a backend for exactly two things: POST /api/signup
# and the seed call that follows it. Everything else, including the whole demo,
# is inlined and works offline.
#
# The legal pages cannot simply be linked from a static build. Their hrefs are
# absolute (/site/legal/...), so publishing site/ as the web root puts index.html
# at / but 404s every legal link, while publishing the repo root fixes the links
# and buries the page at /site/index.html. And a published artifact is a single
# file, where a separate legal/terms.html cannot exist at all. Inlining both
# documents as #terms and #privacy sections resolves every one of those cases at
# once, from the same TERMS_BODY and PRIVACY_BODY the hosted pages use, so there
# is still one source of truth.

STATIC_SIGNUP = """  <h2 class="rv">The hosted workspace opens shortly</h2>
  <p class="lede rv">It is built and tested, and it opens to early users first.
  One message reserves your place.</p>
"""


# Inlining both legal documents as open sections made them roughly half the
# rendered page, which turned a marketing site into a document. They are still
# present in full, in the same file, from the same source of truth; they are
# just folded. <details> was chosen over a <dialog> because it needs no
# JavaScript to open: with scripting off the summary is still a working control.
STATIC_LEGAL_CSS = """
.legalstrip{padding:clamp(40px,5vw,64px) 0 0;border-top:1px solid var(--line-1)}
.legalhead{font-family:var(--font-mono);font-size:10px;letter-spacing:0.18em;
text-transform:uppercase;color:var(--text-tertiary);margin-bottom:16px}
.legalhead span{color:var(--text-tertiary)}
.legaldoc>summary{font-size:15px !important;padding:14px 16px !important;
font-family:var(--font-sans) !important;letter-spacing:0 !important}
.legaldoc>summary::after{content:"Read" !important}
.legaldoc[open]>summary::after{content:"Close" !important}
.cmds{background:var(--surface-sunken);border:1px solid var(--line-1);
border-radius:var(--radius-1);padding:16px 18px;margin-top:14px;max-width:560px;
overflow-x:auto}
.cmds code{font-family:var(--font-mono);font-size:12px;line-height:1.9;
color:var(--text-primary);white-space:pre}
.legalset{display:grid;gap:12px;margin-top:clamp(28px,3.6vw,44px)}
.legaldoc{background:rgba(26,26,29,0.8);border:1px solid var(--line-1);
border-radius:var(--radius-2);overflow:hidden}
.legaldoc>summary{list-style:none;cursor:pointer;padding:clamp(18px,2.2vw,24px);
display:flex;align-items:center;gap:14px;font-family:var(--font-display);
font-size:clamp(18px,1.9vw,23px);letter-spacing:-0.015em}
.legaldoc>summary::-webkit-details-marker{display:none}
.legaldoc>summary::after{content:"Open";margin-left:auto;font-family:var(--font-mono);
font-size:9.5px;letter-spacing:0.17em;text-transform:uppercase;color:var(--text-tertiary);
border:1px solid var(--line-2);border-radius:var(--radius-1);padding:6px 11px}
.legaldoc[open]>summary::after{content:"Close";color:var(--signal);border-color:var(--line-signal)}
.legaldoc>summary:hover{background:rgba(232,232,228,0.03)}
.legaldoc .legalwrap{padding:0 clamp(18px,2.2vw,24px) clamp(20px,2.4vw,28px);
border-top:1px solid var(--line-1);max-width:74ch}
.legaldoc .legalwrap>*:first-child{margin-top:20px}
"""

STATIC_LEGAL_JS = """/* a link to #terms or #privacy opens the folded document it names */
function openLegal(){
  const d=document.getElementById(location.hash.slice(1));
  if(d&&d.tagName==="DETAILS"){d.open=true;d.scrollIntoView({block:"start"})}
}
addEventListener("hashchange",openLegal);openLegal();

"""


def _static_legal_doc(anchor, heading, body):
    return (f'<details class="legaldoc rv" id="{anchor}">'
            f'<summary>{heading}</summary>'
            f'<div class="legalwrap">{DRAFT_NOTE}{body}</div></details>')


def _static_legal_section():
    """A quiet strip above the footer, not a numbered section.

    Legal had a band and a full size heading, which gave it the same billing as
    the product story and ended the page on two grey rows nobody reads. It is
    still here in full and still one tap away; it just no longer competes.
    """
    return ('<section class="legalstrip"><div class="wrap">'
            '<p class="legalhead">The documents, in full. '
            '<span>Founder drafted, pending review by counsel.</span></p>'
            '<div class="legalset">'
            + _static_legal_doc("terms", "Terms of use", TERMS_BODY)
            + _static_legal_doc("privacy", "Privacy and data handling", PRIVACY_BODY)
            + '</div></div></section>')


def build_static(page: str) -> str:
    """Turn the hosted page into a self-contained, backend-free one."""
    import re as _re

    # 1. swap the hosted signup form for the early access block. The region is
    # delimited by explicit markers rather than located by tag index: the
    # section now also carries the contact block, which must survive.
    a, b = "<!--HOSTED-ONLY-START-->", "<!--HOSTED-ONLY-END-->"
    if page.count(a) != 1 or page.count(b) != 1:
        raise SystemExit("expected exactly one hosted-only marker pair")
    start = page.index(a)
    end = page.index(b) + len(b)
    page = page[:start] + STATIC_SIGNUP + page[end:]

    # 2. inline both legal documents, folded, before the footer
    page = page.replace("<footer>", _static_legal_section() + "\n<footer>", 1)

    # 3. point every legal link at those documents
    page = page.replace('href="/site/legal/terms.html"', 'href="#terms"')
    page = page.replace('href="/site/legal/privacy.html"', 'href="#privacy"')

    # 4. the folded documents need one handler, and it has to be inserted above
    #    the signup block so that step 5 does not cut it away again
    marker = "/* \u2500\u2500 signup \u2500"
    # This has bitten once already: a CSS section header used the same words, so
    # replace(count=1) hit the stylesheet and the cut below removed the wrong
    # span. The marker has to be unique in the file or the build is guessing.
    if page.count(marker) != 1:
        raise SystemExit(f"expected exactly one {marker!r}, found {page.count(marker)}")
    page = page.replace(marker, STATIC_LEGAL_JS + marker, 1)

    # 5. drop the JS that talks to the backend. The signup handlers reference
    #    elements that no longer exist, so leaving them throws on load.
    start_js = page.index(marker)
    end_js = page.index("</script>", start_js)
    page = page[:start_js] + page[end_js:]

    # 6. add the legal typography the hosted legal pages carry, plus the
    #    disclosure styling that only the static build needs
    page = page.replace("/* \u2500\u2500 reveal \u2500", LEGAL_CSS + STATIC_LEGAL_CSS
                        + "\n/* \u2500\u2500 reveal \u2500", 1)

    for banned in ("/api/signup", "/site/legal/", "[CALENDAR_LINK]", "[JURISDICTION]"):
        if banned in page:
            raise SystemExit(f"static build still contains {banned!r}")
    return page

def main() -> None:
    store = EvidenceStore("acme", ROOT / "data" / "evidence")
    retriever = Retriever(store)
    out = {}
    for key, label, text, trap in SHOWCASE:
        out[key] = {"label": label, "question": text, "trap": trap,
                    **answer(text, store, retriever)}
    bank = build_bank(store, retriever)
    chain = build_chain(ROOT)
    demo = {"showcase": out, "bank": bank, "chain": chain}

    dest = ROOT / "site" / "demo" / "contracts.json"
    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_text(json.dumps(demo, indent=2, ensure_ascii=False), encoding="utf-8")
    refused = sum(1 for v in out.values() if v["tone"] != "ok")
    print(f"wrote {dest}: {len(out)} showcase contracts ({refused} refusals), "
          f"{len(bank)} bank entries, {len(chain['events'])} chain events")

    # the flagship card is only honest if the cert gate actually fired
    iso_flags = out["iso"]["flags"]
    if not any(f.startswith("CERT_INFERENCE_BLOCKED") for f in iso_flags):
        print(f"  WARNING: the ISO showcase no longer triggers the cert gate "
              f"(flags={iso_flags}). The landing card claims it does.")
    for key, v in out.items():
        print(f"  {key:6} {v['verdict']}")

    # assemble the final landing page: brand CSS + fonts + demo data inlined,
    # so the page is one self-contained file with zero external requests
    from constat import brand
    template = (ROOT / "site" / "index.template.html").read_text(encoding="utf-8")
    page = template.replace("/*@CSS@*/", brand.stylesheet())
    page = page.replace("/*@DEMO@*/", json.dumps(demo, ensure_ascii=False))
    page = page.replace("<!--@CHIPS@-->", chips_html(out))
    page = page.replace("<!--@CHAIN@-->", chain_html(chain))
    page = page.replace("<!--@WORKBOOK@-->", workbook_html())
    page = page.replace("@WBTOTAL@", str(workbook_rowcount()))
    for token, uri in shot_uris().items():
        page = page.replace(f"@{token}@", uri)
    first_key = next(iter(out))
    for name, html in prerender(out[first_key]).items():
        page = page.replace(f"<!--@{name}@-->", html)
    leftovers = [m for m in ("@CSS@", "@DEMO@", "@CHIPS@", "@CHAIN@", "@WORKBOOK@",
                             "@Q0@", "@VERDICT0@",
                             "@ANSWER0@", "@PROV0@", "@GAPS0@", "@SHOT_REPORT@",
                             "@SHOT_REVIEW@", "@SHOT_WORKSPACE@",
                             "@WBTOTAL@") if m in page]
    if leftovers:
        raise SystemExit(f"unfilled template placeholders: {leftovers}")

    # The media tokens are filled per output, because the two builds resolve
    # them differently: the hosted page links the film so the document stays
    # small, the single-file build has no choice but to inline it.
    hosted = page
    for token, uri in media_uris(inline=False).items():
        hosted = hosted.replace(f"@{token}@", uri)
    if "@VIDEO_MP4@" in hosted or "@POSTER_URI@" in hosted:
        raise SystemExit("unfilled media placeholders in the hosted page")
    (ROOT / "site" / "index.html").write_text(hosted, encoding="utf-8")
    print(f"wrote site/index.html ({len(hosted) // 1024}KB)")

    for p in build_legal(ROOT):
        print(f"wrote {p.relative_to(ROOT)} ({p.stat().st_size // 1024}KB)")

    static = build_static(page)
    for token, uri in media_uris(inline=True).items():
        static = static.replace(f"@{token}@", uri)
    if "@VIDEO_MP4@" in static or "@POSTER_URI@" in static:
        raise SystemExit("unfilled media placeholders in the static page")
    static_path = ROOT / "site" / "static" / "index.html"
    static_path.parent.mkdir(parents=True, exist_ok=True)
    static_path.write_text(static, encoding="utf-8")
    print(f"wrote site/static/index.html ({len(static) // 1024}KB, no backend)")


if __name__ == "__main__":
    main()
