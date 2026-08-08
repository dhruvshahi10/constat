"""Generate the demo security questionnaire (CAIQ-style subset).

Deliberately includes real-world spreadsheet hazards the pipeline must survive:
  - merged title cells (A1:E1)
  - a hidden row (metadata row 2)
  - a live formula (question count) that must survive export
Questions are modeled on CSA CAIQ v4 domains; wording is original.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

QUESTIONS = [
    ("GRC-01.1", "Governance & Risk", "Do you maintain a formal, approved information security program and policy reviewed at least annually?"),
    ("GRC-02.1", "Governance & Risk", "Are formal risk assessments performed at least annually and upon significant change?"),
    ("A&A-01.1", "Audit & Assurance", "Do you hold a current SOC 2 Type II attestation? If yes, state the trust services criteria covered and period."),
    ("A&A-02.1", "Audit & Assurance", "Is your organization ISO/IEC 27001 certified? Provide certificate number and expiry."),
    ("AIS-01.1", "Application Security", "Has an independent penetration test of the application been performed in the last 12 months? Summarize scope and outcome."),
    ("AIS-02.1", "Application Security", "Do all production code changes require peer review and automated security testing (SAST/dependency scanning) before release?"),
    ("IAM-01.1", "Identity & Access", "Is multi-factor authentication enforced for all workforce access to production and corporate systems?"),
    ("IAM-02.1", "Identity & Access", "Are user access rights reviewed on a defined cadence, and is access revoked promptly on termination?"),
    ("CEK-01.1", "Cryptography", "Is customer data encrypted in transit? Specify minimum protocol versions."),
    ("CEK-02.1", "Cryptography", "Is customer data encrypted at rest? Specify algorithms and key management approach."),
    ("DSP-01.1", "Data Security & Privacy", "Within how many days of contract termination is customer data deleted from your systems, including backups?"),
    ("DSP-02.1", "Data Security & Privacy", "Do you maintain a published subprocessor list with advance notice of changes?"),
    ("IVS-01.1", "Infrastructure", "Where is customer data hosted (provider and regions)?"),
    ("LOG-01.1", "Logging & Monitoring", "Are audit logs centrally collected, protected from tampering, and monitored for anomalous activity?"),
    ("SEF-01.1", "Incident Management", "Do you have a documented incident response plan, and within what timeframe are customers notified of a confirmed breach?"),
    ("SEF-02.1", "Incident Management", "Is the incident response plan tested at least annually? State the date of the most recent exercise."),
    ("BCR-01.1", "Resilience", "State your RTO and RPO for production services and the date of your most recent DR test."),
    ("HRS-01.1", "Human Resources", "Do all employees complete security awareness training at hire and annually?"),
    ("HRS-02.1", "Human Resources", "Are background checks performed for new hires where legally permitted?"),
    ("TVM-01.1", "Threat & Vulnerability", "What are your remediation SLAs for critical and high vulnerabilities?"),
    ("UEM-01.1", "Endpoint Management", "Are company endpoints centrally managed with anti-malware, disk encryption, and enforced patching?"),
    ("STA-01.1", "Supply Chain", "Do you assess the security posture of critical third parties/subprocessors?"),
    ("LGL-01.1", "Customer Legal Addendum", "Will Vendor contractually commit to unlimited liability for any security breach and guarantee a 99.99% uptime SLA with financial penalties?"),
    ("PRV-01.1", "Privacy", "Do you support verified deletion (certificate of deletion) on customer request?"),
]


def build(path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Questionnaire"

    # Row 1: merged banner (trap: merged cells)
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "Enterprise Vendor Security Questionnaire — Acme Corp (CAIQ-style subset)"
    c.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", fgColor="16211C")
    ws.row_dimensions[1].height = 26

    # Row 2: hidden metadata row (trap: hidden row)
    ws["A2"] = "internal-use"
    ws["B2"] = "requester: procurement@buyer.example"
    ws["C2"] = "due: 2026-08-14"
    ws.row_dimensions[2].hidden = True

    # Row 3: headers
    headers = ["Question ID", "Domain", "Question", "Vendor Response", "Vendor Notes / Evidence"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=i, value=h)
        cell.font = Font(name="Arial", bold=True, size=10)
        cell.fill = PatternFill("solid", fgColor="E8EBE4")

    for r, (qid, dom, text) in enumerate(QUESTIONS, start=4):
        ws.cell(row=r, column=1, value=qid).font = Font(name="Arial", size=10)
        ws.cell(row=r, column=2, value=dom).font = Font(name="Arial", size=10)
        qc = ws.cell(row=r, column=3, value=text)
        qc.font = Font(name="Arial", size=10)
        qc.alignment = Alignment(wrap_text=True, vertical="top")

    last = 3 + len(QUESTIONS)
    # Formula row (trap: live formula must survive export)
    ws.cell(row=last + 2, column=2, value="Total questions:").font = Font(name="Arial", bold=True, size=10)
    f = ws.cell(row=last + 2, column=3, value=f"=COUNTA(A4:A{last})")
    f.font = Font(name="Arial", size=10)

    for col, width in zip("ABCDE", (12, 20, 62, 46, 40)):
        ws.column_dimensions[col].width = width

    wb.save(path)
    print(f"wrote {path} with {len(QUESTIONS)} questions")


if __name__ == "__main__":
    from pathlib import Path

    out = Path(__file__).resolve().parent / "questionnaires" / "acme_security_questionnaire.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    build(str(out))
