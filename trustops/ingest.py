"""Real-document ingestion — the onboarding path for an actual client corpus.

The engine reads markdown with governed frontmatter. Clients have PDFs, Word
policies and spreadsheet registers. This module closes that gap: it extracts
text from what a client actually sends, proposes the governance metadata the
gates depend on, and stages the result for a human to confirm.

Two rules make this safe rather than merely convenient:

  1. Nothing is ever auto-approved. Every staged source lands with
     `approval_status: draft`, which `gates.post_gate` refuses to cite. A named
     human must promote it. The engine will not answer from a document nobody
     has vouched for.
  2. Every inferred field is recorded as inferred. `REVIEW.md` lists exactly
     which values were guessed and which were read from the document, so the
     reviewer knows where to look rather than re-reading everything.

Staging lives in `<tenant>/_staging/`, a subdirectory — `EvidenceStore` globs
`*.md` at the tenant root, so staged material is invisible to retrieval by
construction, not by a filter someone can forget.

Zero new dependencies: PDF via the `pdftotext` binary, DOCX via stdlib zipfile
plus ElementTree, XLSX via the openpyxl already in use.
"""
from __future__ import annotations

import json
import re
import shutil
import subprocess
import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from xml.etree import ElementTree

from .evidence import parse_source

SUPPORTED = {".md", ".markdown", ".txt", ".pdf", ".docx", ".xlsx"}

W_NS = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"

# --- source typing -----------------------------------------------------------
# Filename first, body second: a retention standard that mentions a "certificate
# of deletion" is a standard, not a certificate. Roadmap is checked before
# everything else — a document describing a future state is never evidence of a
# current one, whatever standard it names.
#
# `certificate` and `attestation` are deliberately ABSENT from this table.
# Those two types are the only ones that satisfy the certification gate
# (`gates.CERT_EVIDENCE_TYPES`), so assigning one is the single most
# consequential inference this module could make. It refuses to make it: a
# likely certificate is flagged for a human to type deliberately.
TYPE_RULES: list[tuple[str, str, re.Pattern]] = [
    ("roadmap", "RMP", re.compile(
        r"\b(roadmap|road\s*map|target\s*state|future\s*state|intends?\s+to\s+achieve|"
        r"planned\s+(?:certification|audit)|gap\s*analysis)\b", re.I)),
    ("report", "RPT", re.compile(
        r"\b(report|assessment|penetration\s*test|pentest|audit\s*result|scan\s*result|"
        r"findings)\b", re.I)),
    ("standard", "STD", re.compile(
        r"\b(standard|baseline|hardening|configuration\s*guide|specification)\b", re.I)),
    ("register", "REG", re.compile(
        r"\b(register|inventory|matrix|tracker|schedule\b|list\b)\b", re.I)),
    ("plan", "PLN", re.compile(
        r"\b(plan|playbook|runbook|procedure|programme)\b", re.I)),
    ("policy", "POL", re.compile(r"\b(policy|policies|charter)\b", re.I)),
]
DEFAULT_TYPE, DEFAULT_PREFIX = "policy", "POL"

# Matched against the filename and opening text only to RAISE A FLAG — never to
# set a type. See TYPE_RULES above.
CERT_HINT = re.compile(
    r"\b(certificate|certification|certified|attestation|attested|soc\s*[123]\b|"
    r"type\s*(?:i{1,2}|[12])\b|bridge\s*letter|accredit\w*)\b", re.I)

TOPIC_VOCAB: dict[str, re.Pattern] = {
    "encryption": re.compile(r"\b(encrypt\w*|tls|aes|kms|key management|cipher)\b", re.I),
    "access control": re.compile(r"\b(access control|least privilege|rbac|provisioning|authorisation|authorization)\b", re.I),
    "authentication": re.compile(r"\b(multi-?factor|mfa|2fa|sso|password|authentication)\b", re.I),
    "data retention": re.compile(r"\b(retention|deletion|erasure|purge|disposal)\b", re.I),
    "incident response": re.compile(r"\b(incident response|breach|security incident|escalation)\b", re.I),
    "business continuity": re.compile(r"\b(business continuity|disaster recovery|rto|rpo|failover)\b", re.I),
    "secure development": re.compile(r"\b(sdlc|code review|sast|dast|dependency scan|secure development)\b", re.I),
    "human resources": re.compile(r"\b(background check|onboarding|security awareness|training|offboarding)\b", re.I),
    "penetration testing": re.compile(r"\b(penetration test|pentest|red team|ethical hack)\b", re.I),
    "subprocessors": re.compile(r"\b(subprocessor|sub-processor|third part\w+|vendor management|supply chain)\b", re.I),
    "logging": re.compile(r"\b(audit log|logging|siem|monitoring|telemetry)\b", re.I),
    "vulnerability management": re.compile(r"\b(vulnerabilit\w+|patch\w*|remediation sla|cve)\b", re.I),
    "endpoint": re.compile(r"\b(endpoint|laptop|mdm|anti-?malware|disk encryption)\b", re.I),
    "privacy": re.compile(r"\b(privacy|gdpr|personal data|dpa\b|data subject)\b", re.I),
    "ai governance": re.compile(r"\b(artificial intelligence|\bai\b|machine learning|llm|model governance|iso\s*42001)\b", re.I),
    "risk management": re.compile(r"\b(risk assessment|risk register|risk treatment|threat model)\b", re.I),
    "governance": re.compile(r"\b(isms|governance|management review|internal audit)\b", re.I),
}

# Machine-checkable assertions. These drive deterministic contradiction
# detection, so they are proposed conservatively and always flagged for review.
ASSERTION_RULES: list[tuple[str, re.Pattern]] = [
    ("customer_data_deletion_days",
     re.compile(r"(?:delet\w+|eras\w+|purg\w+|destroy\w+)[^.\n]{0,80}?\bwithin\s+(\d{1,4})\s+days?", re.I)),
    ("breach_notification_hours",
     re.compile(r"(?:notif\w+|inform\w+)[^.\n]{0,80}?\bwithin\s+(\d{1,3})\s+hours?", re.I)),
    ("rto_hours", re.compile(r"\brto\b[^.\n]{0,40}?(\d{1,3})\s*hours?", re.I)),
    ("rpo_hours", re.compile(r"\brpo\b[^.\n]{0,40}?(\d{1,3})\s*hours?", re.I)),
    ("access_review_days",
     re.compile(r"access (?:rights? )?(?:are )?review\w*[^.\n]{0,60}?every\s+(\d{1,3})\s+days?", re.I)),
    ("critical_vuln_remediation_days",
     re.compile(r"critical[^.\n]{0,60}?\bwithin\s+(\d{1,3})\s+days?", re.I)),
]

_VERSION_PAT = re.compile(r"\bv(?:ersion)?[\s:.]*(\d+(?:\.\d+){0,2})\b", re.I)
_EMAIL_PAT = re.compile(r"[\w.+-]+@[\w-]+\.[\w.-]+")
_ISO_DATE = r"(\d{4}-\d{2}-\d{2})"
_LONG_DATE = r"(\d{1,2}\s+[A-Z][a-z]+\s+\d{4}|[A-Z][a-z]+\s+\d{1,2},?\s+\d{4})"
_EFFECTIVE_PAT = re.compile(
    rf"\b(?:effective|issued|approved|published|adopted)(?:\s+(?:date|on))?\s*[:\-]?\s*(?:{_ISO_DATE}|{_LONG_DATE})", re.I)
_EXPIRY_PAT = re.compile(
    rf"\b(?:expir\w*|valid\s+(?:un)?til|next\s+review|review\s+(?:date|due)|renewal)\s*[:\-]?\s*(?:{_ISO_DATE}|{_LONG_DATE})", re.I)
_TITLE_SKIP = re.compile(r"^(confidential|internal|page\s+\d|draft)\b", re.I)

UNCONFIRMED_OWNER = "UNCONFIRMED@confirm-before-approval.invalid"


# --- text extraction ---------------------------------------------------------
class ExtractionError(RuntimeError):
    pass


def _pdf_text(path: Path) -> str:
    exe = shutil.which("pdftotext")
    if exe:
        proc = subprocess.run([exe, "-layout", "-enc", "UTF-8", str(path), "-"],
                              capture_output=True, timeout=120)
        if proc.returncode == 0 and proc.stdout.strip():
            return proc.stdout.decode("utf-8", errors="replace")
    try:  # optional, only if the operator installed it
        from pypdf import PdfReader  # type: ignore
    except ImportError as exc:
        raise ExtractionError(
            f"{path.name}: cannot read PDF. Install poppler (`brew install poppler`) "
            f"for the `pdftotext` binary, or `pip install pypdf`."
        ) from exc
    return "\n\n".join(page.extract_text() or "" for page in PdfReader(str(path)).pages)


def _docx_text(path: Path) -> str:
    with zipfile.ZipFile(path) as zf:
        xml = zf.read("word/document.xml")
    root = ElementTree.fromstring(xml)
    paras = []
    for p in root.iter(f"{W_NS}p"):
        text = "".join(t.text or "" for t in p.iter(f"{W_NS}t"))
        if text.strip():
            paras.append(text.strip())
    return "\n\n".join(paras)


def _xlsx_text(path: Path) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    out = []
    for ws in wb.worksheets:
        out.append(f"## {ws.title}")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
            if cells:
                out.append(" | ".join(cells))
    return "\n\n".join(out)


def extract_text(path: Path) -> str:
    """Plain text from whatever the client actually sent."""
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix not in SUPPORTED:
        raise ExtractionError(f"{path.name}: unsupported type '{suffix}' "
                              f"(supported: {', '.join(sorted(SUPPORTED))})")
    if suffix == ".pdf":
        text = _pdf_text(path)
    elif suffix == ".docx":
        text = _docx_text(path)
    elif suffix == ".xlsx":
        text = _xlsx_text(path)
    else:
        text = path.read_text(encoding="utf-8", errors="replace")
    if not text.strip():
        raise ExtractionError(f"{path.name}: no extractable text "
                              f"(scanned image PDF? needs OCR before ingestion)")
    return text


# --- metadata proposal -------------------------------------------------------
def _normalise_body(text: str) -> str:
    # strip existing frontmatter so re-staging an engine-format file is idempotent
    text = re.sub(r"^---\n.*?\n---\n", "", text, flags=re.DOTALL)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+\n", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def _parse_any_date(raw: str) -> date | None:
    raw = raw.strip().rstrip(",")
    for fmt in ("%Y-%m-%d", "%d %B %Y", "%B %d, %Y", "%B %d %Y", "%d %b %Y", "%b %d, %Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def _first_date(pat: re.Pattern, text: str) -> date | None:
    m = pat.search(text)
    if not m:
        return None
    for group in m.groups():
        if group:
            parsed = _parse_any_date(group)
            if parsed:
                return parsed
    return None


def _propose_title(text: str, path: Path) -> str:
    for line in text.splitlines()[:25]:
        line = line.strip().lstrip("#").strip()
        if 6 <= len(line) <= 90 and not _TITLE_SKIP.match(line) and not line.endswith("."):
            return line
    return path.stem.replace("_", " ").replace("-", " ").strip().title()


def _propose_type(path: Path, text: str) -> tuple[str, str, bool]:
    """(type, id prefix, read_from_filename). Filename evidence beats body prose."""
    for type_name, prefix, pat in TYPE_RULES:
        if pat.search(path.stem):
            return type_name, prefix, True
    for type_name, prefix, pat in TYPE_RULES:
        if pat.search(text[:1500]):
            return type_name, prefix, False
    return DEFAULT_TYPE, DEFAULT_PREFIX, False


def _slug(path: Path) -> str:
    words = re.findall(r"[A-Za-z0-9]+", path.stem.upper())
    drop = {"THE", "A", "AN", "OF", "AND", "FOR", "V", "FINAL", "DRAFT", "COPY", "SIGNED"}
    words = [w for w in words if w not in drop and not re.fullmatch(r"\d{1,2}", w)]
    return "-".join(words[:2]) or "SOURCE"


@dataclass
class Proposal:
    """A staged source plus an explicit account of what was inferred."""
    source_id: str
    path: Path                       # the original client file
    meta: dict[str, str]
    assertions: dict[str, str]
    body: str
    inferred: list[str] = field(default_factory=list)   # fields a human must confirm
    read_from_document: list[str] = field(default_factory=list)

    def to_markdown(self) -> str:
        lines = ["---"]
        for key in ("source_id", "title", "type", "version", "effective_date",
                    "expiry_date", "owner", "approval_status", "topics"):
            lines.append(f"{key}: {self.meta[key]}")
        for key, value in self.assertions.items():
            lines.append(f"assert.{key}: {value}")
        lines += [f"ingested_from: {self.path.name}",
                  f"ingested_at: {datetime.now(timezone.utc).date().isoformat()}",
                  f"inferred_fields: {', '.join(self.inferred) or 'none'}",
                  "---", "", self.body, ""]
        return "\n".join(lines)


def propose_source(path: Path, text: str, existing_ids: set[str] | None = None) -> Proposal:
    """Propose governance metadata for one extracted document.

    Anything not stated in the document is inferred and named as inferred. The
    two fields that matter most to the gates — `approval_status` and
    `expiry_date` — never silently favour citability: status is always `draft`,
    and an unstated expiry defaults to one year from the effective date and is
    flagged.
    """
    existing_ids = existing_ids or set()
    body = _normalise_body(text)
    inferred: list[str] = []
    read: list[str] = []

    type_name, prefix, from_filename = _propose_type(path, body)
    (read if from_filename else inferred).append("type")
    if CERT_HINT.search(f"{path.stem} {body[:800]}"):
        # The certification gate accepts only type certificate|attestation. This
        # module will not hand that key over on a keyword match.
        inferred.append("type — POSSIBLE CERTIFICATE/ATTESTATION, typed as "
                        f"'{type_name}' pending human confirmation")

    base = f"{prefix}-{_slug(path)}"
    source_id, n = base, 1
    while source_id in existing_ids:
        n += 1
        source_id = f"{base}-{n:03d}"

    version_match = _VERSION_PAT.search(f"{path.stem} {body[:800]}")
    if version_match:
        version, _ = version_match.group(1), read.append("version")
    else:
        version = "0.1"
        inferred.append("version")

    effective = _first_date(_EFFECTIVE_PAT, body)
    if effective:
        read.append("effective_date")
    else:
        effective = datetime.fromtimestamp(path.stat().st_mtime).date()
        inferred.append("effective_date (file modified time)")

    expiry = _first_date(_EXPIRY_PAT, body)
    if expiry:
        read.append("expiry_date")
    else:
        expiry = effective + timedelta(days=365)
        inferred.append("expiry_date (effective + 12 months)")

    email = _EMAIL_PAT.search(body)
    if email:
        owner, _ = email.group(0), read.append("owner")
    else:
        owner = UNCONFIRMED_OWNER
        inferred.append("owner")

    topics = [name for name, pat in TOPIC_VOCAB.items() if pat.search(body)]
    if topics:
        read.append("topics")
    else:
        inferred.append("topics (none matched)")

    assertions: dict[str, str] = {}
    for key, pat in ASSERTION_RULES:
        m = pat.search(body)
        if m:
            assertions[key] = m.group(1)
    if assertions:
        inferred.append(f"assertions ({', '.join(assertions)})")

    meta = {
        "source_id": source_id,
        "title": _propose_title(body, path),
        "type": type_name,
        "version": version,
        "effective_date": effective.isoformat(),
        "expiry_date": expiry.isoformat(),
        "owner": owner,
        # never auto-approved: the gates refuse to cite a draft source
        "approval_status": "draft",
        "topics": ", ".join(topics),
    }
    return Proposal(source_id=source_id, path=path, meta=meta, assertions=assertions,
                    body=body, inferred=inferred, read_from_document=read)


# --- staging -----------------------------------------------------------------
@dataclass
class StagingResult:
    tenant: str
    staged: list[Proposal]
    failed: list[tuple[Path, str]]
    staging_dir: Path
    review_path: Path


def staging_dir(evidence_root: Path, tenant: str) -> Path:
    return Path(evidence_root) / tenant / "_staging"


def existing_source_ids(evidence_root: Path, tenant: str) -> set[str]:
    ids: set[str] = set()
    tenant_dir = Path(evidence_root) / tenant
    for folder in (tenant_dir, staging_dir(evidence_root, tenant)):
        if not folder.is_dir():
            continue
        for p in folder.glob("*.md"):
            m = re.search(r"^source_id:\s*(\S+)", p.read_text(encoding="utf-8"), re.M)
            if m:
                ids.add(m.group(1))
    return ids


def stage_corpus(tenant: str, src_dir: Path, evidence_root: Path) -> StagingResult:
    """Extract, propose and stage every supported document in a folder."""
    src_dir, evidence_root = Path(src_dir), Path(evidence_root)
    out_dir = staging_dir(evidence_root, tenant)
    out_dir.mkdir(parents=True, exist_ok=True)
    ids = existing_source_ids(evidence_root, tenant)

    staged: list[Proposal] = []
    failed: list[tuple[Path, str]] = []
    candidates = sorted(p for p in src_dir.rglob("*")
                        if p.is_file() and p.suffix.lower() in SUPPORTED
                        and not p.name.startswith((".", "~$")))
    for path in candidates:
        try:
            proposal = propose_source(path, extract_text(path), ids)
        except (ExtractionError, OSError, zipfile.BadZipFile, ElementTree.ParseError) as exc:
            failed.append((path, str(exc)))
            continue
        target = out_dir / f"{proposal.source_id}.md"
        target.write_text(proposal.to_markdown(), encoding="utf-8")
        try:   # a staged file that will not parse is a staging bug, not a review item
            parse_source(target, tenant)
        except (ValueError, KeyError) as exc:
            target.unlink(missing_ok=True)
            failed.append((path, f"proposed frontmatter did not parse: {exc}"))
            continue
        ids.add(proposal.source_id)
        staged.append(proposal)

    review_path = out_dir / "REVIEW.md"
    review_path.write_text(_review_markdown(tenant, staged, failed, src_dir), encoding="utf-8")
    return StagingResult(tenant=tenant, staged=staged, failed=failed,
                         staging_dir=out_dir, review_path=review_path)


def _review_markdown(tenant: str, staged: list[Proposal],
                     failed: list[tuple[Path, str]], src_dir: Path) -> str:
    today = datetime.now(timezone.utc).date().isoformat()
    lines = [
        f"# Evidence review queue — {tenant}",
        "",
        f"Staged {today} from `{src_dir}`. **Nothing here is citable yet.**",
        "Every staged source carries `approval_status: draft`, and the citation gate",
        "refuses to cite a source that is not approved. Confirm the fields below, then",
        "promote with a named approver:",
        "",
        "```bash",
        f"python onboard.py promote --tenant {tenant} --id <SOURCE-ID> \\",
        '    --actor "Your Name <you@company.com>"',
        "```",
        "",
        f"## Staged sources ({len(staged)})",
        "",
    ]
    for p in staged:
        lines += [
            f"### `{p.source_id}` — {p.meta['title']}",
            f"- source file: `{p.path.name}`",
            f"- type **{p.meta['type']}** · version **{p.meta['version']}** · "
            f"in force **{p.meta['effective_date']} → {p.meta['expiry_date']}**",
            f"- owner: `{p.meta['owner']}`",
            f"- topics: {p.meta['topics'] or '_none detected_'}",
        ]
        if p.assertions:
            lines.append("- machine-checked assertions (drive contradiction detection — "
                         "confirm the numbers): " +
                         ", ".join(f"`{k}={v}`" for k, v in p.assertions.items()))
        lines.append(f"- **confirm before approving:** "
                     f"{', '.join(p.inferred) if p.inferred else '_nothing inferred_'}")
        if p.meta["owner"] == UNCONFIRMED_OWNER:
            lines.append("- ⚠ no owner found in the document — routing on failure will "
                         "have nobody to name until this is set")
        lines.append("")
    if failed:
        lines += [f"## Could not ingest ({len(failed)})", ""]
        lines += [f"- `{path.name}` — {reason}" for path, reason in failed] + [""]
    return "\n".join(lines)


# --- promotion ---------------------------------------------------------------
def corpus_log_path(evidence_root: Path, tenant: str) -> Path:
    return Path(evidence_root) / tenant / "_corpus_log.jsonl"


def promote(tenant: str, source_id: str, evidence_root: Path, actor: str,
            approve: bool = True) -> Path:
    """Move a staged source into the live corpus under a named human's approval.

    Approval is an act with an author. The corpus log records who vouched for
    the document and when; the source itself is only marked approved because a
    person said so, never because ingestion succeeded.
    """
    evidence_root = Path(evidence_root)
    staged = staging_dir(evidence_root, tenant) / f"{source_id}.md"
    if not staged.is_file():
        raise FileNotFoundError(f"no staged source '{source_id}' for tenant '{tenant}'")
    if not actor or "@" not in actor:
        raise ValueError("promotion requires a named approver, e.g. "
                         '--actor "Priya Nair <priya@company.com>"')
    text = staged.read_text(encoding="utf-8")
    if approve:
        text = re.sub(r"^approval_status:.*$", "approval_status: approved", text,
                      count=1, flags=re.M)
        text = re.sub(r"^(inferred_fields:.*)$",
                      rf"\1\napproved_by: {actor}", text, count=1, flags=re.M)
    target = evidence_root / tenant / f"{source_id}.md"
    target.write_text(text, encoding="utf-8")
    parse_source(target, tenant)          # must parse in the live corpus, or it does not ship
    staged.unlink()
    with corpus_log_path(evidence_root, tenant).open("a", encoding="utf-8") as f:
        f.write(json.dumps({
            "ts": datetime.now(timezone.utc).isoformat(),
            "actor": actor,
            "action": "APPROVED_SOURCE" if approve else "PROMOTED_AS_DRAFT",
            "source_id": source_id,
        }, ensure_ascii=False) + "\n")
    return target
