"""Generate a questionnaire workbook from a question list.

Hosted runs answer a fixed question set with no customer XLSX. Rather than
teaching the pipeline a second input path, we build a real workbook and feed
the exact code path the constitution certifies (ingest -> gates -> export).
The generated file deliberately keeps the structural hazards T6 guards:
merged banner row, hidden metadata row, header row 3, and a live formula.
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def build_questionnaire_workbook(questions: list[tuple[str, str, str]],
                                 dst: Path, org: str) -> Path:
    """questions: (question_id, domain, text). Returns dst."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Questionnaire"

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = f"Security Questionnaire, {org} (TrustOps hosted run)"
    c.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", fgColor="16211C")
    ws.row_dimensions[1].height = 26

    ws["A2"] = "internal-use"
    ws["B2"] = f"generated: {date.today().isoformat()}"
    ws["C2"] = "origin: trustops-hosted"
    ws.row_dimensions[2].hidden = True

    headers = ["Question ID", "Domain", "Question", "Vendor Response", "Vendor Notes / Evidence"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=i, value=h)
        cell.font = Font(name="Arial", bold=True, size=10)
        cell.fill = PatternFill("solid", fgColor="E8EBE4")

    for r, (qid, dom, text) in enumerate(questions, start=4):
        ws.cell(row=r, column=1, value=qid).font = Font(name="Arial", size=10)
        ws.cell(row=r, column=2, value=dom).font = Font(name="Arial", size=10)
        qc = ws.cell(row=r, column=3, value=text)
        qc.font = Font(name="Arial", size=10)
        qc.alignment = Alignment(wrap_text=True, vertical="top")

    last = 3 + len(questions)
    ws.cell(row=last + 2, column=2, value="Total questions:").font = \
        Font(name="Arial", bold=True, size=10)
    ws.cell(row=last + 2, column=3, value=f"=COUNTA(A4:A{last})").font = \
        Font(name="Arial", size=10)

    for col, width in zip("ABCDE", (12, 20, 62, 46, 40)):
        ws.column_dimensions[col].width = width

    dst.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dst)
    return dst
