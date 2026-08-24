"""Ingest and structure-preserving export.

Contract (blueprint p.10, definition of done):
  - import a real XLSX and preserve identifiers and order on export
  - write ONLY the response/notes columns; question text is never modified
  - merged cells, hidden rows, and formulas survive the round trip

Row identity is carried on the Question object itself (q.row), so export is a
positional write-back, not a fuzzy match.

Layout is DETECTED, not assumed. Real buyer questionnaires — CAIQ, SIG Lite,
bespoke vendor workbooks — put their headers on different rows and in different
column orders, so a hardcoded layout silently ingests zero questions from a
real client's file. `detect_layout` resolves the header row and column roles
from header semantics; the demo workbook resolves to the historical constants
(header row 3, columns 1-5), so the structural round-trip guarantee is
unchanged.
"""
from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

from .models import Draft, Question

# Historical constants: the demo workbook's layout. Retained as the documented
# fallback and as the shape detection must reproduce for that file.
HEADER_ROW = 3
COL_ID, COL_DOMAIN, COL_Q, COL_RESP, COL_NOTES = 1, 2, 3, 4, 5

MAX_HEADER_SCAN = 15          # real headers live near the top; deeper is a footer
MAX_HEADER_CELL_CHARS = 60    # a 74-char banner is a title, not a column header
RESP_HEADER = "Vendor Response"
NOTES_HEADER = "Vendor Notes / Evidence"

# Header role patterns. Applied in precedence order, first match wins per cell,
# so "Question ID" resolves to id (not question) and "Vendor Response" to resp.
_ID_PAT = re.compile(r"\b(id|ids|ref|no|num|number|item|q#|#)\b", re.I)
_RESP_PAT = re.compile(r"\b(response|answer|answers|reply|disposition)\b", re.I)
_NOTES_PAT = re.compile(
    r"\b(note|notes|comment|comments|evidence|remark|remarks|reference|references|"
    r"attachment|attachments|justification)\b", re.I)
_DOMAIN_PAT = re.compile(
    r"\b(domain|category|categories|section|family|group|area|topic)\b", re.I)
_Q_PAT = re.compile(
    r"\b(question|questions|description|control|controls|requirement|requirements|"
    r"inquiry|criteria|criterion|ask|prompt)\b", re.I)

_ROLE_ORDER = (("id", _ID_PAT), ("resp", _RESP_PAT), ("notes", _NOTES_PAT),
               ("domain", _DOMAIN_PAT), ("q", _Q_PAT))


@dataclass
class Layout:
    """Resolved position of every column role the engine reads or writes."""
    header_row: int
    col_q: int
    col_resp: int
    col_notes: int
    col_id: int | None = None
    col_domain: int | None = None
    created_resp: bool = False    # column did not exist; export must add a header
    created_notes: bool = False

    def as_dict(self) -> dict:
        return {"header_row": self.header_row, "col_id": self.col_id,
                "col_domain": self.col_domain, "col_q": self.col_q,
                "col_resp": self.col_resp, "col_notes": self.col_notes,
                "created_resp": self.created_resp, "created_notes": self.created_notes}


def _cell_role(value) -> str | None:
    if not isinstance(value, str):
        return None
    text = value.strip()
    if not text or len(text) > MAX_HEADER_CELL_CHARS:
        return None
    for role, pat in _ROLE_ORDER:
        if pat.search(text):
            return role
    return None


def _row_roles(ws, row: int, max_col: int) -> dict[str, int]:
    roles: dict[str, int] = {}
    for col in range(1, max_col + 1):
        role = _cell_role(ws.cell(row=row, column=col).value)
        if role and role not in roles:   # leftmost wins for a duplicated role
            roles[role] = col
    return roles


def detect_layout(ws) -> Layout:
    """Resolve the header row and column roles of an arbitrary questionnaire sheet.

    A candidate header row must expose a question column. Rows are scored by how
    many distinct roles they carry, so a header beats a stray data row that
    happens to contain the word "control". Missing response/notes columns are
    appended rather than overwriting an existing column.
    """
    max_col = ws.max_column or 1
    best: tuple[int, int, dict[str, int]] | None = None
    for row in range(1, min(MAX_HEADER_SCAN, ws.max_row or 1) + 1):
        roles = _row_roles(ws, row, max_col)
        if "q" not in roles:
            continue
        score = 2 + sum(1 for r in ("id", "domain", "resp", "notes") if r in roles)
        if best is None or score > best[0]:
            best = (score, row, roles)

    if best is None:
        # No recognisable header: fall back to the documented demo layout rather
        # than guessing. Ingest will simply find no questions and say so.
        return Layout(header_row=HEADER_ROW, col_q=COL_Q, col_resp=COL_RESP,
                      col_notes=COL_NOTES, col_id=COL_ID, col_domain=COL_DOMAIN)

    _, header_row, roles = best
    next_free = max_col + 1
    col_resp = roles.get("resp")
    created_resp = col_resp is None
    if created_resp:
        col_resp = next_free
        next_free += 1
    col_notes = roles.get("notes")
    created_notes = col_notes is None
    if created_notes:
        col_notes = next_free
    return Layout(header_row=header_row, col_q=roles["q"], col_resp=col_resp,
                  col_notes=col_notes, col_id=roles.get("id"),
                  col_domain=roles.get("domain"), created_resp=created_resp,
                  created_notes=created_notes)


def _is_data_row(qtext, qid_value, layout: Layout) -> bool:
    """A data row has question text that is not a formula, and — when the sheet
    has an ID column — a string identifier. That second rule is what keeps the
    demo workbook's `=COUNTA(...)` footer out of the question list."""
    if qtext is None:
        return False
    text = str(qtext).strip()
    if not text or text.startswith("="):
        return False
    if layout.col_id is not None:
        return bool(qid_value) and isinstance(qid_value, str)
    return True


# --- XLSX --------------------------------------------------------------------
def _ingest_xlsx(path: Path, layout: Layout | None) -> list[Question]:
    wb = load_workbook(path)
    ws = wb.active
    lay = layout or detect_layout(ws)
    out: list[Question] = []
    for row in range(lay.header_row + 1, ws.max_row + 1):
        qtext = ws.cell(row=row, column=lay.col_q).value
        qid_value = ws.cell(row=row, column=lay.col_id).value if lay.col_id else None
        if not _is_data_row(qtext, qid_value, lay):
            continue
        qid = qid_value.strip() if isinstance(qid_value, str) else f"R{row}"
        domain = (str(ws.cell(row=row, column=lay.col_domain).value or "").strip()
                  if lay.col_domain else "")
        out.append(Question(question_id=qid, row=row, domain=domain,
                            text=str(qtext).strip()))
    return out


def _export_xlsx(src: Path, dst: Path, questions: list[Question],
                 drafts: dict[str, Draft], layout: Layout | None) -> None:
    wb = load_workbook(src)   # default load: formulas preserved as formulas
    ws = wb.active
    lay = layout or detect_layout(ws)
    body = Font(name="Arial", size=10)
    wrap = Alignment(wrap_text=True, vertical="top")
    header_font = Font(name="Arial", bold=True, size=10)
    if lay.created_resp:
        ws.cell(row=lay.header_row, column=lay.col_resp, value=RESP_HEADER).font = header_font
    if lay.created_notes:
        ws.cell(row=lay.header_row, column=lay.col_notes, value=NOTES_HEADER).font = header_font
    for q in questions:
        d = drafts[q.question_id]
        # write ONLY the response columns; never touch the id or question columns
        rc = ws.cell(row=q.row, column=lay.col_resp, value=_response_text(d))
        nc = ws.cell(row=q.row, column=lay.col_notes, value=_notes_text(d))
        rc.font = body; rc.alignment = wrap
        nc.font = body; nc.alignment = wrap
    wb.save(dst)


# --- CSV ---------------------------------------------------------------------
def _csv_rows(path: Path) -> list[list[str]]:
    with path.open(newline="", encoding="utf-8-sig") as f:
        return [row for row in csv.reader(f)]


class _CsvSheet:
    """Minimal duck-typed sheet so CSV reuses the same detection logic."""

    def __init__(self, rows: list[list[str]]):
        self._rows = rows
        self.max_row = len(rows)
        self.max_column = max((len(r) for r in rows), default=1)

    def cell(self, row: int, column: int):
        class _C:
            value = None
        c = _C()
        if 1 <= row <= len(self._rows):
            r = self._rows[row - 1]
            if 1 <= column <= len(r):
                c.value = r[column - 1]
        return c


def _ingest_csv(path: Path, layout: Layout | None) -> list[Question]:
    sheet = _CsvSheet(_csv_rows(path))
    lay = layout or detect_layout(sheet)
    out: list[Question] = []
    for row in range(lay.header_row + 1, sheet.max_row + 1):
        qtext = sheet.cell(row=row, column=lay.col_q).value
        qid_value = sheet.cell(row=row, column=lay.col_id).value if lay.col_id else None
        if not _is_data_row(qtext, qid_value, lay):
            continue
        qid = qid_value.strip() if isinstance(qid_value, str) and qid_value.strip() else f"R{row}"
        domain = (str(sheet.cell(row=row, column=lay.col_domain).value or "").strip()
                  if lay.col_domain else "")
        out.append(Question(question_id=qid, row=row, domain=domain, text=str(qtext).strip()))
    return out


def _export_csv(src: Path, dst: Path, questions: list[Question],
                drafts: dict[str, Draft], layout: Layout | None) -> None:
    rows = _csv_rows(src)
    lay = layout or detect_layout(_CsvSheet(rows))
    width = max(lay.col_resp, lay.col_notes, max((len(r) for r in rows), default=1))
    rows = [r + [""] * (width - len(r)) for r in rows]
    if lay.created_resp:
        rows[lay.header_row - 1][lay.col_resp - 1] = RESP_HEADER
    if lay.created_notes:
        rows[lay.header_row - 1][lay.col_notes - 1] = NOTES_HEADER
    for q in questions:
        d = drafts[q.question_id]
        rows[q.row - 1][lay.col_resp - 1] = _response_text(d)
        rows[q.row - 1][lay.col_notes - 1] = _notes_text(d)
    with dst.open("w", newline="", encoding="utf-8") as f:
        csv.writer(f).writerows(rows)


# --- public API --------------------------------------------------------------
def ingest_questionnaire(path: Path, layout: Layout | None = None) -> list[Question]:
    path = Path(path)
    if path.suffix.lower() == ".csv":
        return _ingest_csv(path, layout)
    return _ingest_xlsx(path, layout)


def layout_of(path: Path) -> Layout:
    """Resolved layout for a questionnaire — used by onboarding to show an
    operator what the engine believes the file's shape is before a run."""
    path = Path(path)
    if path.suffix.lower() == ".csv":
        return detect_layout(_CsvSheet(_csv_rows(path)))
    return detect_layout(load_workbook(path).active)


def _response_text(d: Draft) -> str:
    if d.answer:
        return d.answer
    if d.route == "LEGAL":
        return "[ROUTED TO LEGAL] Contractual commitment — not drafted by the answer engine."
    return "[ABSTAINED] Insufficient approved evidence — see exception notes."


def _notes_text(d: Draft) -> str:
    parts = []
    if d.citations:
        parts.append("Evidence: " + "; ".join(
            f"{c.source_id} v{c.version} ({c.location})" for c in d.citations))
    if d.gaps:
        parts.append("Gaps: " + " | ".join(d.gaps))
    if d.route:
        parts.append(f"Routed: {d.route}")
    parts.append(f"[status={d.status.label} coverage={d.evidence_coverage.value} "
                 f"risk={d.risk.value} "
                 f"human_review={'yes' if d.requires_human else 'approved'}]")
    return "\n".join(parts)


def export_answers(src: Path, dst: Path, questions: list[Question],
                   drafts: dict[str, Draft], layout: Layout | None = None) -> None:
    src, dst = Path(src), Path(dst)
    if src.suffix.lower() == ".csv":
        _export_csv(src, dst, questions, drafts, layout)
    else:
        _export_xlsx(src, dst, questions, drafts, layout)
